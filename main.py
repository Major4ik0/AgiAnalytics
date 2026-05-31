# -*- coding: utf-8 -*-
import sys
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QGridLayout, QLabel,
                             QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
                             QMessageBox, QComboBox, QFrame,
                             QGroupBox, QTabWidget, QDialog, QDialogButtonBox,
                             QFormLayout,
                             QFileDialog, QToolBar, QStatusBar,
                             QScrollArea, QAction, QCheckBox, QProgressDialog,
                             QSizePolicy, QInputDialog, QListWidget)
from PyQt5.QtCore import Qt, pyqtSignal, QThread
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor
from datetime import datetime
from database import Database
from resource_helper import get_icon_path, resource_path
from statistics_widget import StatisticsWidget
import pandas as pd
import os

os.environ['QT_MAC_WANTS_LAYER'] = '1'

ICONS = get_icon_path("icon.ico") or "icons/icon.ico"


class LoginWindow(QWidget):
    """Окно авторизации"""
    login_success = pyqtSignal(dict)  # Сигнал успешной авторизации

    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon(ICONS))
        self.setWindowTitle('Авторизация')
        self.setFixedSize(400, 300)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Заголовок
        title = QLabel('AgiAnalytics')
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_font = QFont()
        title_font.setPointSize(18)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # Поля ввода
        form_widget = QWidget()
        form_layout = QFormLayout()

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText('Введите логин')
        self.username_input.setMinimumHeight(40)
        self.username_input.returnPressed.connect(self.authenticate)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText('Введите пароль')
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setMinimumHeight(40)
        self.password_input.returnPressed.connect(self.authenticate)

        form_layout.addRow('Логин:', self.username_input)
        form_layout.addRow('Пароль:', self.password_input)
        form_widget.setLayout(form_layout)
        layout.addWidget(form_widget)

        # Кнопка входа
        self.login_btn = QPushButton('Войти')
        self.login_btn.setMinimumHeight(45)
        self.login_btn.clicked.connect(self.authenticate)
        layout.addWidget(self.login_btn)

        # Информация
        info_label = QLabel()
        info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_label.setStyleSheet('color: #666; font-style: italic;')
        layout.addWidget(info_label)
        self.setLayout(layout)

    def keyPressEvent(self, event):
        """Обработка нажатия клавиш"""
        if event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter:
            self.authenticate()
        else:
            super().keyPressEvent(event)

    def authenticate(self):
        username = self.username_input.text().strip()
        password = self.password_input.text().strip()

        if not username or not password:
            QMessageBox.warning(self, 'Ошибка', 'Заполните все поля!')
            return

        db = Database()
        user = db.get_user_by_credentials(username, password)

        if user:
            user_dict = dict(user)
            self.login_success.emit(user_dict)
        else:
            QMessageBox.critical(self, 'Ошибка', 'Неверный логин или пароль!')


class ApplicantDialog(QDialog):
    """Диалог добавления/редактирования абитуриента"""

    def __init__(self, applicant_data=None, user_role='user', db=None, parent=None):
        super().__init__(parent)
        self.applicant_data = applicant_data
        self.user_role = user_role
        self.db = db
        self.setModal(True)

        if applicant_data:
            self.setWindowTitle('Редактировать абитуриента')
        else:
            self.setWindowTitle('Добавить абитуриента')

        self.setMinimumSize(700, 750)
        self.setMaximumSize(900, 800)
        self.init_ui()

    def init_ui(self):
        """Инициализация интерфейса"""
        # Основной layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        # Создаем скролл область для длинной формы
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
        """)

        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(20)

        # ========== БЛОК 1: ИНФОРМАЦИЯ ОБ АБИТУРИЕНТЕ ==========
        applicant_group = QGroupBox("Информация об абитуриенте")
        applicant_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #3498db;
            }
        """)

        applicant_layout = QFormLayout()
        applicant_layout.setSpacing(12)
        applicant_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # ФИО абитуриента (обязательное)
        self.applicant_name = QLineEdit()
        self.applicant_name.setPlaceholderText("Иванов Иван Иванович")
        self.applicant_name.setMinimumHeight(35)
        applicant_layout.addRow("ФИО абитуриента *:", self.applicant_name)

        # Субъект РФ
        self.region = QComboBox()
        self.region.setEditable(True)
        self.region.setPlaceholderText("Выберите или введите субъект РФ")
        self.load_regions() # Загружаем регионы из БД
        self.region.setMinimumHeight(35)
        applicant_layout.addRow("Субъект РФ:", self.region)

        # Населенный пункт
        self.city = QLineEdit()
        self.city.setPlaceholderText("Населенный пункт")
        self.city.setMinimumHeight(35)
        applicant_layout.addRow("Населенный пункт:", self.city)

        # Категория
        self.category = QComboBox()
        self.category.addItems(["м", "ж", "всл"])
        self.category.setMinimumHeight(35)
        # Добавляем подсказки
        self.category.setItemText(0, "м (Мужчина)")
        self.category.setItemText(1, "ж (Женщина)")
        self.category.setItemText(2, "всл (Военнослужащий)")
        applicant_layout.addRow("Категория *:", self.category)

        # Телефон с маской
        self.phone = QLineEdit()
        self.phone.setPlaceholderText("+7 (999) 999-99-99")
        self.phone.setInputMask("+7 (999) 999-99-99;_")
        self.phone.setMinimumHeight(35)
        self.phone.textChanged.connect(self.validate_phone)
        applicant_layout.addRow("Телефон *:", self.phone)

        # Образование
        self.education = QComboBox()
        self.education.setMinimumHeight(35)
        self.load_education_types()
        applicant_layout.addRow("Образование:", self.education)

        # Статус
        self.status = QComboBox()
        self.status.addItems(["поступает", "отказывается"])
        self.status.setMinimumHeight(35)
        # Добавляем иконки в текст
        self.status.setItemText(0, "Поступает")
        self.status.setItemText(1, "Отказывается")
        applicant_layout.addRow("Статус *:", self.status)

        # Статус документов
        self.document_status = QComboBox()
        self.document_status.setMinimumHeight(35)
        self.load_document_statuses()
        applicant_layout.addRow("Документы:", self.document_status)

        applicant_group.setLayout(applicant_layout)
        scroll_layout.addWidget(applicant_group)

        # ========== БЛОК 2: ИНФОРМАЦИЯ ОБ АГИТАТОРЕ ==========
        agitator_group = QGroupBox("Информация об агитаторе")
        agitator_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #2ecc71;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #2ecc71;
            }
        """)

        agitator_layout = QFormLayout()
        agitator_layout.setSpacing(12)
        agitator_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # Тип агитатора
        self.agitator_type_widget = QWidget()
        agitator_type_layout = QHBoxLayout(self.agitator_type_widget)
        agitator_type_layout.setContentsMargins(0, 0, 0, 0)
        agitator_type_layout.setSpacing(15)

        self.agitator_is_cadet = QCheckBox("Агитатор - курсант")
        self.agitator_is_cadet.setStyleSheet("""
            QCheckBox {
                spacing: 8px;
                font-weight: bold;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)
        self.agitator_is_cadet.stateChanged.connect(self.on_agitator_type_changed)

        self.agitator_is_officer = QCheckBox("Агитатор - офицер/военнослужащий")
        self.agitator_is_officer.setStyleSheet("""
            QCheckBox {
                spacing: 8px;
                font-weight: bold;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)
        self.agitator_is_officer.stateChanged.connect(self.on_agitator_type_changed)

        agitator_type_layout.addWidget(self.agitator_is_cadet)
        agitator_type_layout.addWidget(self.agitator_is_officer)
        agitator_type_layout.addStretch()
        agitator_layout.addRow("Тип агитатора:", self.agitator_type_widget)

        # Подразделение
        self.agitator_department = QComboBox()
        self.agitator_department.currentTextChanged.connect(self.on_department_changed)
        self.agitator_department.setEditable(True)
        self.agitator_department.setMinimumHeight(35)
        self.load_departments()
        agitator_layout.addRow("Подразделение:", self.agitator_department)

        # ФИО агитатора
        self.agitator_name = QLineEdit()
        self.agitator_name.setPlaceholderText("Фамилия И.О.")
        self.agitator_name.setMinimumHeight(35)
        agitator_layout.addRow("ФИО агитатора *:", self.agitator_name)

        # Контейнер для полей курсанта (по умолчанию скрыт)
        self.cadet_widget = QWidget()
        cadet_layout = QGridLayout(self.cadet_widget)
        cadet_layout.setContentsMargins(0, 0, 0, 0)
        cadet_layout.setSpacing(10)

        # Курс (для курсанта)
        self.agitator_course = QComboBox()
        self.agitator_course.addItems(["1 курс", "2 курс", "3 курс", "4 курс", "5 курс"])
        self.agitator_course.setMinimumHeight(35)
        cadet_layout.addWidget(QLabel("Курс:"), 0, 0)
        cadet_layout.addWidget(self.agitator_course, 0, 1)

        # Группа (для курсанта)
        self.agitator_group = QLineEdit()
        self.agitator_group.setPlaceholderText("Номер группы")
        self.agitator_group.setMinimumHeight(35)
        cadet_layout.addWidget(QLabel("Группа:"), 1, 0)
        cadet_layout.addWidget(self.agitator_group, 1, 1)

        agitator_layout.addRow("", self.cadet_widget)

        # Контейнер для полей офицера/военнослужащего (по умолчанию скрыт)
        self.officer_widget = QWidget()
        officer_layout = QHBoxLayout(self.officer_widget)
        officer_layout.setContentsMargins(0, 0, 0, 0)

        # Звание (для офицера/военнослужащего)
        self.agitator_rank = QComboBox()
        self.agitator_rank.addItems([
            "ряд.", "ефр.", "мл. серж.", "серж.", "ст. серж.", "старшина",
            "прапорщик", "ст. прапорщик", "мл. лейт.", "лейтенант", "ст. лейтенант",
            "капитан", "майор", "подполковник", "полковник", "генерал-майор",
            "генерал-лейтенант", "генерал-полковник"
        ])
        self.agitator_rank.setMinimumHeight(35)
        officer_layout.addWidget(QLabel("Звание:"))
        officer_layout.addWidget(self.agitator_rank)
        officer_layout.addStretch()

        agitator_layout.addRow("", self.officer_widget)

        agitator_group.setLayout(agitator_layout)
        scroll_layout.addWidget(agitator_group)

        # ========== БЛОК 3: ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ ==========
        extra_group = QGroupBox("Дополнительная информация")
        extra_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #e67e22;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #e67e22;
            }
        """)

        extra_layout = QFormLayout()
        extra_layout.setSpacing(12)

        # Примечания
        self.notes = QLineEdit()
        self.notes.setPlaceholderText("Дополнительные замечания, комментарии...")
        self.notes.setMinimumHeight(35)
        extra_layout.addRow("Примечания:", self.notes)

        extra_group.setLayout(extra_layout)
        scroll_layout.addWidget(extra_group)

        # Информация об обязательных полях
        info_label = QLabel("* - поля, обязательные для заполнения")
        info_label.setStyleSheet("color: #e74c3c; font-size: 11px; margin-top: 5px;")
        scroll_layout.addWidget(info_label)

        scroll_layout.addStretch()
        scroll_area.setWidget(scroll_widget)
        main_layout.addWidget(scroll_area)

        # Кнопки
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)

        # Стилизация кнопок
        ok_button = button_box.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("Сохранить")
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)

        cancel_button = button_box.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)

        main_layout.addWidget(button_box)

        # Заполнение данных если редактирование
        if self.applicant_data:
            self.load_applicant_data()

        # Устанавливаем начальное состояние полей
        self.on_agitator_type_changed()

    def load_education_types(self):
        """Загрузка типов образования из БД"""
        if self.db:
            education_list = self.db.get_education_types()
            self.education.clear()
            self.education.addItems(education_list)
        else:
            # Данные по умолчанию
            self.education.addItems(["СОШ", "СПО", "СВУ", "ПКУ", "КК"])

    def load_document_statuses(self):
        """Загрузка статусов документов из БД"""
        if self.db:
            statuses = self.db.get_document_statuses()
            self.document_status.clear()
            self.document_status.addItems(statuses)
        else:
            # Данные по умолчанию
            self.document_status.addItems(["ВК", "ОК", "ВА ВКО"])

    def load_departments(self):
        """Загрузка подразделений из БД (без дублей)"""
        if self.db:
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT DISTINCT name FROM departments WHERE type != "root" ORDER BY name')
            departments = [row['name'] for row in cursor.fetchall()]

            self.agitator_department.clear()
            self.agitator_department.addItem("")
            for dept_name in departments:
                self.agitator_department.addItem(dept_name)

    def load_regions(self):
        """Загрузка регионов из БД"""
        if self.db:
            regions = self.db.get_regions()
            self.region.clear()
            self.region.addItem("Выберите или введите субъект РФ")
            self.region.addItems(regions)
        else:
            # Данные по умолчанию
            self.region.addItems([
                "Москва", "Санкт-Петербург", "Московская область", "Ленинградская область",
                "Краснодарский край", "Красноярский край", "Ставропольский край",
                "Республика Татарстан", "Республика Башкортостан", "Свердловская область",
                "Ростовская область", "Новосибирская область", "Челябинская область",
                "Нижегородская область", "Самарская область", "Омская область",
                "Воронежская область", "Пермский край", "Волгоградская область"
            ])

    def on_agitator_type_changed(self):
        """Обработка изменения типа агитатора"""
        is_cadet = self.agitator_is_cadet.isChecked()
        is_officer = self.agitator_is_officer.isChecked()

        # Получаем выбранное подразделение
        department = self.agitator_department.currentText()

        # Проверяем, может ли быть курсант в этом подразделении
        can_be_cadet = 'Факультет 1' in department or 'Факультет 2' in department

        # Обновляем состояние чекбокса курсанта
        self.agitator_is_cadet.setEnabled(can_be_cadet)

        if not can_be_cadet and is_cadet:
            # Если выбрано неподходящее подразделение, переключаем на офицера
            self.agitator_is_cadet.setChecked(False)
            self.agitator_is_officer.setChecked(True)
            is_cadet = False
            is_officer = True

        # Показываем/скрываем соответствующие поля
        self.cadet_widget.setVisible(is_cadet and can_be_cadet)
        self.officer_widget.setVisible(is_officer)

        # Если ни один не выбран, показываем оба (по умолчанию курсант если можно)
        if not is_cadet and not is_officer:
            if can_be_cadet:
                self.agitator_is_cadet.setChecked(True)
            else:
                self.agitator_is_officer.setChecked(True)

    def on_department_changed(self, text):
        """Обработка изменения подразделения"""
        self.on_agitator_type_changed()

    def validate_phone(self, text):
        """Валидация номера телефона"""
        # Удаляем маску для проверки
        clean_text = text.replace('(', '').replace(')', '').replace('-', '').replace(' ', '').replace('_', '').replace(
            '+', '')

        if len(clean_text) < 11:
            self.phone.setStyleSheet("border: 1px solid #e74c3c;")
        else:
            self.phone.setStyleSheet("border: 1px solid #2ecc71;")

    def validate_and_accept(self):
        """Валидация обязательных полей (только те, что в таблице)"""
        errors = []

        # Обязательные поля из таблицы:
        # 1. ФИО абитуриента
        if not self.applicant_name.text().strip():
            errors.append("ФИО абитуриента")
            self.applicant_name.setStyleSheet("border: 2px solid #e74c3c;")
        else:
            self.applicant_name.setStyleSheet("")

        # 2. Телефон (должен быть заполнен)
        phone_clean = self.clean_phone_number(self.phone.text())
        if not phone_clean or len(phone_clean) < 10:
            errors.append("Телефон")
            self.phone.setStyleSheet("border: 2px solid #e74c3c;")
        else:
            self.phone.setStyleSheet("")

        # 3. ФИО агитатора (обязательно)
        if not self.agitator_name.text().strip():
            errors.append("ФИО агитатора")
            self.agitator_name.setStyleSheet("border: 2px solid #e74c3c;")
        else:
            self.agitator_name.setStyleSheet("")

        # 4. Подразделение (обязательно)
        if not self.agitator_department.currentText().strip():
            errors.append("Подразделение")
            self.agitator_department.setStyleSheet("border: 2px solid #e74c3c;")
        else:
            self.agitator_department.setStyleSheet("")

        # Категория - всегда выбрана по умолчанию, не проверяем
        # Статус - всегда выбран по умолчанию, не проверяем

        # Для курсанта: если выбран тип "курсант", то группа обязательна
        if self.agitator_is_cadet.isChecked():
            if not self.agitator_group.text().strip():
                errors.append("Группа (для курсанта)")
                self.agitator_group.setStyleSheet("border: 2px solid #e74c3c;")
            else:
                self.agitator_group.setStyleSheet("")

        # Если есть ошибки
        if errors:
            error_msg = "Пожалуйста, заполните следующие обязательные поля:\n• " + "\n• ".join(errors)
            QMessageBox.warning(self, "Ошибка валидации", error_msg)
            return

        self.accept()

    def clean_phone_number(self, phone):
        """Очистка номера телефона от форматирования"""
        if not phone:
            return ""

        # Удаляем все нецифровые символы
        digits = ''.join(filter(str.isdigit, phone))

        if not digits:
            return ""

        # Приводим к формату 7XXXXXXXXXX
        if digits.startswith('8') and len(digits) == 11:
            return '7' + digits[1:]
        elif len(digits) == 10:
            return '7' + digits
        elif digits.startswith('7') and len(digits) == 11:
            return digits

        return digits

    def load_applicant_data(self):
        """Загрузка данных абитуриента для редактирования"""
        # Абитуриент
        self.applicant_name.setText(self.applicant_data.get('applicant_name', ''))

        # Регион
        region = self.applicant_data.get('region', '')
        if region:
            index = self.region.findText(region)
            if index >= 0:
                self.region.setCurrentIndex(index)
            else:
                self.region.setEditText(region)

        self.city.setText(self.applicant_data.get('city', ''))

        # Категория
        category = self.applicant_data.get('category', 'м')
        if category == 'м':
            self.category.setCurrentIndex(0)
        elif category == 'ж':
            self.category.setCurrentIndex(1)
        else:
            self.category.setCurrentIndex(2)

        # Телефон
        phone = self.applicant_data.get('phone', '')
        if phone:
            formatted = self.format_phone_for_display(phone)
            self.phone.setText(formatted)

        # Образование
        edu = self.applicant_data.get('education', '')
        if edu:
            index = self.education.findText(edu)
            if index >= 0:
                self.education.setCurrentIndex(index)

        # Статус
        status = self.applicant_data.get('status', 'поступает')
        self.status.setCurrentIndex(0 if status == 'поступает' else 1)

        # Документы
        doc_status = self.applicant_data.get('document_status', '')
        if doc_status:
            index = self.document_status.findText(doc_status)
            if index >= 0:
                self.document_status.setCurrentIndex(index)

        # Агитатор
        department = self.applicant_data.get('agitator_department', '')
        if department:
            index = self.agitator_department.findText(department)
            if index >= 0:
                self.agitator_department.setCurrentIndex(index)
            else:
                self.agitator_department.setEditText(department)

        self.agitator_name.setText(self.applicant_data.get('agitator_name', ''))

        # Тип агитатора
        is_cadet = self.applicant_data.get('agitator_is_cadet', False)
        if is_cadet:
            self.agitator_is_cadet.setChecked(True)
            course = self.applicant_data.get('agitator_course', '1 курс')
            index = self.agitator_course.findText(course)
            if index >= 0:
                self.agitator_course.setCurrentIndex(index)
            self.agitator_group.setText(self.applicant_data.get('agitator_group', ''))
        else:
            self.agitator_is_officer.setChecked(True)
            rank = self.applicant_data.get('agitator_rank', '')
            if rank:
                index = self.agitator_rank.findText(rank)
                if index >= 0:
                    self.agitator_rank.setCurrentIndex(index)

        # Примечания
        self.notes.setText(self.applicant_data.get('notes', ''))

        # Обновляем состояние полей
        self.on_agitator_type_changed()

    @staticmethod
    def format_phone_for_display(phone):
        """Форматирование телефона для отображения в маске"""
        digits = ''.join(filter(str.isdigit, str(phone)))

        if len(digits) >= 11:
            # Формат +7 (XXX) XXX-XX-XX
            return f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
        return phone

    def get_data(self):
        """Получение данных из формы"""
        # Получаем чистый телефон
        phone = self.clean_phone_number(self.phone.text())

        # Получаем категорию
        category_map = {0: 'м', 1: 'ж', 2: 'всл'}
        category = category_map.get(self.category.currentIndex(), 'м')

        # Получаем статус
        status = 'поступает' if self.status.currentIndex() == 0 else 'отказывается'

        # Данные абитуриента
        data = {
            'applicant_name': self.applicant_name.text().strip(),
            'region': self.region.currentText().strip() if self.region.currentText() != "Выберите или введите субъект РФ" else "",
            'city': self.city.text().strip(),
            'category': category,
            'phone': phone,
            'education': self.education.currentText(),
            'status': status,
            'document_status': self.document_status.currentText(),
            'agitator_department': self.agitator_department.currentText().strip(),
            'agitator_name': self.agitator_name.text().strip(),
            'agitator_is_cadet': self.agitator_is_cadet.isChecked(),
            'notes': self.notes.text().strip(),
        }

        # Данные в зависимости от типа агитатора
        if self.agitator_is_cadet.isChecked():
            data['agitator_course'] = self.agitator_course.currentText()
            data['agitator_group'] = self.agitator_group.text().strip()
            data['agitator_rank'] = ''
        else:
            data['agitator_course'] = ''
            data['agitator_group'] = ''
            data['agitator_rank'] = self.agitator_rank.currentText()

        return data


class AdvancedSearchDialog(QDialog):
    """Диалог расширенного поиска"""

    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.filters = {}
        self.setModal(True)
        self.setWindowTitle('Расширенный поиск')
        self.setMinimumSize(500, 450)
        self.setMaximumSize(700, 600)
        self.init_ui()

    def init_ui(self):
        """Инициализация интерфейса"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Заголовок
        title_label = QLabel("Расширенный поиск абитуриентов")
        title_font = title_label.font()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(title_label)

        # Создаем скролл область
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; }")

        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(15)

        # ========== БЛОК 1: ИНФОРМАЦИЯ ОБ АБИТУРИЕНТЕ ==========
        applicant_group = QGroupBox("Информация об абитуриенте")
        applicant_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)

        applicant_layout = QFormLayout()
        applicant_layout.setSpacing(10)

        # ФИО абитуриента
        self.applicant_name = QLineEdit()
        self.applicant_name.setPlaceholderText("Введите ФИО полностью или частично")
        self.applicant_name.setClearButtonEnabled(True)
        applicant_layout.addRow("ФИО абитуриента:", self.applicant_name)

        # Субъект РФ
        self.region = QComboBox()
        self.region.setEditable(True)
        self.load_regions()
        applicant_layout.addRow("Субъект РФ:", self.region)

        # Населенный пункт
        self.city = QLineEdit()
        self.city.setPlaceholderText("Введите населенный пункт")
        self.city.setClearButtonEnabled(True)
        applicant_layout.addRow("Населенный пункт:", self.city)

        # Категория
        self.category = QComboBox()
        self.category.addItems(["все", "м", "ж", "всл"])
        self.category.setItemText(0, "Все категории")
        self.category.setItemText(1, "Мужчина")
        self.category.setItemText(2, "Женщина")
        self.category.setItemText(3, "Военнослужащий")
        applicant_layout.addRow("Категория:", self.category)

        # Образование (множественный выбор)
        self.education_group = QGroupBox("Образование")
        self.education_group.setStyleSheet("QGroupBox { border: none; margin-top: 5px; }")
        education_layout = QHBoxLayout(self.education_group)

        self.education_checkboxes = []
        education_types = self.get_education_types()
        for edu in education_types:
            checkbox = QCheckBox(edu)
            self.education_checkboxes.append(checkbox)
            education_layout.addWidget(checkbox)
        education_layout.addStretch()

        applicant_layout.addRow("", self.education_group)

        # Статус
        self.status = QComboBox()
        self.status.addItems(["все", "поступает", "отказывается"])
        self.status.setItemText(0, "Все статусы")
        applicant_layout.addRow("Статус:", self.status)

        applicant_group.setLayout(applicant_layout)
        scroll_layout.addWidget(applicant_group)

        # ========== БЛОК 2: ИНФОРМАЦИЯ ОБ АГИТАТОРЕ ==========
        agitator_group = QGroupBox("Информация об агитаторе")
        agitator_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)

        agitator_layout = QFormLayout()
        agitator_layout.setSpacing(10)

        # ФИО агитатора
        self.agitator_name = QLineEdit()
        self.agitator_name.setPlaceholderText("Введите ФИО агитатора")
        self.agitator_name.setClearButtonEnabled(True)
        agitator_layout.addRow("ФИО агитатора:", self.agitator_name)

        # Подразделение
        self.agitator_department = QComboBox()
        self.agitator_department.setEditable(True)
        self.load_departments()
        agitator_layout.addRow("Подразделение:", self.agitator_department)

        # Тип агитатора
        self.agitator_type = QComboBox()
        self.agitator_type.addItems(["все", "курсант", "офицер/военнослужащий"])
        self.agitator_type.setItemText(0, "Все")
        agitator_layout.addRow("Тип агитатора:", self.agitator_type)

        agitator_group.setLayout(agitator_layout)
        scroll_layout.addWidget(agitator_group)

        # ========== БЛОК 3: ДОПОЛНИТЕЛЬНЫЕ ПАРАМЕТРЫ ==========
        extra_group = QGroupBox("Дополнительные параметры")
        extra_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)

        extra_layout = QGridLayout()
        extra_layout.setSpacing(10)

        # Статус документов
        extra_layout.addWidget(QLabel("Документы:"), 0, 0)
        self.document_status = QComboBox()
        self.document_status.setEditable(True)
        self.load_document_statuses()
        extra_layout.addWidget(self.document_status, 0, 1)

        # Курс (для курсанта)
        extra_layout.addWidget(QLabel("Курс:"), 1, 0)
        self.course = QComboBox()
        self.course.addItems(["все", "1 курс", "2 курс", "3 курс", "4 курс", "5 курс"])
        self.course.setItemText(0, "Все курсы")
        extra_layout.addWidget(self.course, 1, 1)

        # Группа
        extra_layout.addWidget(QLabel("Группа:"), 2, 0)
        self.group = QLineEdit()
        self.group.setPlaceholderText("Номер группы")
        self.group.setClearButtonEnabled(True)
        extra_layout.addWidget(self.group, 2, 1)

        extra_group.setLayout(extra_layout)
        scroll_layout.addWidget(extra_group)

        # Кнопки быстрого выбора
        quick_buttons_widget = QWidget()
        quick_buttons_layout = QHBoxLayout(quick_buttons_widget)

        self.clear_all_btn = QPushButton("Очистить все")
        self.clear_all_btn.clicked.connect(self.clear_all_filters)

        self.select_none_btn = QPushButton("Снять все выделения")
        self.select_none_btn.clicked.connect(self.clear_education_selection)

        quick_buttons_layout.addWidget(self.clear_all_btn)
        quick_buttons_layout.addWidget(self.select_none_btn)
        quick_buttons_layout.addStretch()

        scroll_layout.addWidget(quick_buttons_widget)
        scroll_layout.addStretch()

        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # Кнопки диалога
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )

        ok_button = button_box.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("Найти")
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)

        cancel_button = button_box.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")

        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        layout.addWidget(button_box)


    def load_regions(self):
        """Загрузка субъектов РФ (уникальные из БД)"""
        if self.db:
            regions = self.db.get_regions()
            self.region.clear()
            self.region.addItem("Все регионы") # Добавляем опцию "Все регионы"
            self.region.addItems(regions)
        else:
            # Данные по умолчанию
            default_regions = [
                "Москва", "Санкт-Петербург", "Московская область", "Ленинградская область",
                "Краснодарский край", "Красноярский край", "Ставропольский край"
            ]
            self.region.addItems(default_regions)

    def load_departments(self):
        """Загрузка подразделений (уникальные из БД)"""
        if self.db:
            cursor = self.db.conn.cursor()
            cursor.execute(
                'SELECT DISTINCT agitator_department FROM applicants WHERE agitator_department IS NOT NULL AND agitator_department != "" ORDER BY agitator_department')
            departments = [row['agitator_department'] for row in cursor.fetchall()]
            self.agitator_department.clear()
            self.agitator_department.addItem("")
            self.agitator_department.addItems(departments)
        else:
            self.agitator_department.addItems(["", "Факультет 1", "Факультет 2", "Кафедра 1", "Кафедра 2"])

    def get_education_types(self):
        """Получение типов образования"""
        if self.db:
            return self.db.get_education_types()
        return ["СОШ", "СПО", "СВУ", "ПКУ", "КК"]

    def load_document_statuses(self):
        """Загрузка статусов документов"""
        if self.db:
            statuses = self.db.get_document_statuses()
            self.document_status.addItems([""] + statuses)
        else:
            self.document_status.addItems(["", "ВК", "ОК", "ВА ВКО"])

    def clear_all_filters(self):
        """Очистка всех фильтров"""
        self.applicant_name.clear()
        self.region.setCurrentIndex(0)
        self.city.clear()
        self.category.setCurrentIndex(0)
        self.status.setCurrentIndex(0)
        self.agitator_name.clear()
        self.agitator_department.setCurrentIndex(0)
        self.agitator_type.setCurrentIndex(0)
        self.document_status.setCurrentIndex(0)
        self.course.setCurrentIndex(0)
        self.group.clear()
        self.clear_education_selection()

    def clear_education_selection(self):
        """Снятие всех выделений образования"""
        for checkbox in self.education_checkboxes:
            checkbox.setChecked(False)

    def get_filters(self):
        """Получение выбранных фильтров (только заполненные)"""
        filters = {}
        # Абитуриент - только если не пусто
        applicant_name = self.applicant_name.text().strip()
        if applicant_name:
            filters['applicant_name'] = applicant_name
        region = self.region.currentText().strip()
        if region and region != "Все регионы": # Изменено
            filters['region'] = region
        city = self.city.text().strip()
        if city:
            filters['city'] = city
        # Категория - только если не "все"
        category = self.category.currentText()
        if category != "все" and category != "Все категории":
            if category == "Мужчина":
                filters['category'] = 'м'
            elif category == "Женщина":
                filters['category'] = 'ж'
            elif category == "Военнослужащий":
                filters['category'] = 'всл'
        # Образование (только выбранные)
        selected_education = [cb.text() for cb in self.education_checkboxes if cb.isChecked()]
        if selected_education:
            filters['education'] = selected_education
        # Статус - только если не "все"
        status = self.status.currentText()
        if status != "все" and status != "Все статусы":
            filters['status'] = status
        # Агитатор
        agitator_name = self.agitator_name.text().strip()
        if agitator_name:
            filters['agitator_name'] = agitator_name
        agitator_department = self.agitator_department.currentText().strip()
        if agitator_department:
            filters['agitator_department'] = agitator_department
        # Тип агитатора
        agitator_type = self.agitator_type.currentText()
        if agitator_type != "все" and agitator_type != "Все":
            filters['agitator_is_cadet'] = (agitator_type == "курсант")
        # Документы
        document_status = self.document_status.currentText().strip()
        if document_status:
            filters['document_status'] = document_status
        # Курс - только если не "все"
        course = self.course.currentText()
        if course != "все" and course != "Все курсы":
            filters['agitator_course'] = course
        # Группа
        group = self.group.text().strip()
        if group:
            filters['agitator_group'] = group
        return filters


class SelectionStatsDialog(QDialog):
    """Диалог со статистикой по выделенным строкам"""

    def __init__(self, selected_data, parent=None):
        super().__init__(parent)
        self.selected_data = selected_data
        self.setModal(False)  # Не модальный, чтобы можно было продолжать работу
        self.setWindowTitle('Статистика по выделенным записям')
        self.setFixedSize(500, 600)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Заголовок
        title_label = QLabel(f"Статистика по выделенным записям")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50; margin-bottom: 15px;")
        layout.addWidget(title_label)

        # Информация о количестве
        count_label = QLabel(f"Выделено записей: {len(self.selected_data)}")
        count_label.setStyleSheet("font-weight: bold; font-size: 13px; margin-bottom: 10px;")
        layout.addWidget(count_label)

        # Статистика
        stats_group = QGroupBox("Статистика по выделенным записям")
        stats_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)

        stats_layout = QFormLayout()
        stats_layout.setSpacing(12)

        # Вычисляем статистику
        stats = self.calculate_stats()

        # Добавляем статистику
        stats_layout.addRow("Всего:", QLabel(str(stats['total'])))
        stats_layout.addRow("Поступают:", QLabel(f"{stats['applying']} ({stats['applying_percent']:.1f}%)"))
        stats_layout.addRow("Отказываются:", QLabel(f"{stats['refused']} ({stats['refused_percent']:.1f}%)"))

        # Разделитель
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        stats_layout.addRow(line)

        stats_layout.addRow("Мужчины:", QLabel(f"{stats['male']} ({stats['male_percent']:.1f}%)"))
        stats_layout.addRow("Женщины:", QLabel(f"{stats['female']} ({stats['female_percent']:.1f}%)"))
        stats_layout.addRow("Военнослужащие:", QLabel(f"{stats['military']} ({stats['military_percent']:.1f}%)"))

        # Разделитель
        line2 = QFrame()
        line2.setFrameShape(QFrame.Shape.HLine)
        line2.setFrameShadow(QFrame.Shadow.Sunken)
        stats_layout.addRow(line2)

        stats_layout.addRow("ВК:", QLabel(str(stats['doc1'])))
        stats_layout.addRow("ОК:", QLabel(str(stats['doc2'])))
        stats_layout.addRow("ВА ВКО:", QLabel(str(stats['doc3'])))

        # Статистика по курсам (если есть разные курсы)
        if len(stats['courses']) > 1:
            line3 = QFrame()
            line3.setFrameShape(QFrame.Shape.HLine)
            line3.setFrameShadow(QFrame.Shadow.Sunken)
            stats_layout.addRow(line3)

            courses_text = ", ".join([f"{k}: {v}" for k, v in stats['courses'].items()])
            stats_layout.addRow("По курсам агитатора:", QLabel(courses_text))

        stats_group.setLayout(stats_layout)
        layout.addWidget(stats_group)

        # Кнопка копирования
        copy_btn = QPushButton("Копировать статистику")
        copy_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-weight: bold;
                margin-top: 15px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        copy_btn.clicked.connect(self.copy_stats)
        layout.addWidget(copy_btn)

        # Кнопка закрытия
        close_btn = QPushButton("Закрыть")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px;
                margin-top: 5px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)

        self.setLayout(layout)

    def calculate_stats(self):
        """Вычисление статистики по выделенным данным (обновленный для новой структуры)"""
        stats = {
            'total': len(self.selected_data),
            'applying': 0,
            'refused': 0,
            'male': 0,
            'female': 0,
            'military': 0,
            'doc1': 0,  # ВК
            'doc2': 0,  # ОК
            'doc3': 0,  # ВА ВКО
            'courses': {},
            'applying_percent': 0,
            'refused_percent': 0,
            'male_percent': 0,
            'female_percent': 0,
            'military_percent': 0
        }

        for row_data in self.selected_data:
            # Статус поступления
            status = row_data.get('status', '')
            if status == 'поступает':
                stats['applying'] += 1
            elif status == 'отказывается':
                stats['refused'] += 1

            # Категория
            category = row_data.get('category', '')
            if category == 'м':
                stats['male'] += 1
            elif category == 'ж':
                stats['female'] += 1
            elif category == 'всл':
                stats['military'] += 1

            # Статус документов
            doc_status = row_data.get('document_status', '')
            if doc_status == 'ВК':
                stats['doc1'] += 1
            elif doc_status == 'ОК':
                stats['doc2'] += 1
            elif doc_status == 'ВА ВКО':
                stats['doc3'] += 1

            # Курс агитатора
            course = row_data.get('agitator_course', '')
            if course:
                stats['courses'][course] = stats['courses'].get(course, 0) + 1

        # Вычисляем проценты
        total = stats['total']
        if total > 0:
            stats['applying_percent'] = (stats['applying'] / total) * 100
            stats['refused_percent'] = (stats['refused'] / total) * 100
            stats['male_percent'] = (stats['male'] / total) * 100
            stats['female_percent'] = (stats['female'] / total) * 100
            stats['military_percent'] = (stats['military'] / total) * 100

        return stats

    def copy_stats(self):
        """Копирование статистики в буфер обмена"""
        stats = self.calculate_stats()

        clipboard_text = f"""
    СТАТИСТИКА ПО ВЫДЕЛЕННЫМ ЗАПИСЯМ
    {'=' * 40}

    Выделено записей: {stats['total']}

    СТАТУС ПОСТУПЛЕНИЯ:
      • Поступают: {stats['applying']} ({stats['applying_percent']:.1f}%)
      • Отказываются: {stats['refused']} ({stats['refused_percent']:.1f}%)

    КАТЕГОРИИ:
      • Мужчины: {stats['male']} ({stats['male_percent']:.1f}%)
      • Женщины: {stats['female']} ({stats['female_percent']:.1f}%)
      • Военнослужащие: {stats['military']} ({stats['military_percent']:.1f}%)

    СТАТУС ДОКУМЕНТОВ:
      • ВК: {stats['doc1']}
      • ОК: {stats['doc2']}
      • ВА ВКО: {stats['doc3']}

    РАСПРЕДЕЛЕНИЕ ПО КУРСАМ АГИТАТОРА:
    """
        for course, count in sorted(stats['courses'].items()):
            percent = (count / stats['total']) * 100 if stats['total'] > 0 else 0
            clipboard_text += f"  • {course}: {count} ({percent:.1f}%)\n"

        clipboard_text += f"\n{'=' * 40}\n"
        clipboard_text += f"{datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"

        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_text)

        QMessageBox.information(self, "Успех", "Статистика скопирована в буфер обмена!")


class UserDialog(QDialog):
    """Диалог добавления/редактирования пользователя"""

    def __init__(self, user_data=None, db=None, parent=None):
        super().__init__(parent)
        self.user_data = user_data
        self.db = db
        self.setModal(True)

        if user_data:
            self.setWindowTitle('Редактировать пользователя')
        else:
            self.setWindowTitle('Добавить пользователя')

        self.setMinimumSize(600, 750)
        self.setMaximumSize(800, 850)
        self.init_ui()

    def init_ui(self):
        """Инициализация интерфейса"""
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        # Заголовок
        title_label = QLabel("Управление пользователем")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("color: #2c3e50; margin-bottom: 10px;")
        main_layout.addWidget(title_label)

        # Скролл область
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; }")

        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(20)

        # ========== БЛОК 1: ОСНОВНАЯ ИНФОРМАЦИЯ ==========
        main_group = QGroupBox("Основная информация")
        main_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #3498db;
            }
        """)

        main_form = QFormLayout()
        main_form.setSpacing(12)
        main_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # Логин
        self.username = QLineEdit()
        self.username.setPlaceholderText("Введите логин")
        self.username.setMinimumHeight(35)
        main_form.addRow("Логин *:", self.username)

        # Пароль
        self.password = QLineEdit()
        self.password.setPlaceholderText("Введите пароль")
        self.password.setEchoMode(QLineEdit.EchoMode.Password)
        self.password.setMinimumHeight(35)
        main_form.addRow("Пароль *:", self.password)

        # Подтверждение пароля (только для нового пользователя)
        if not self.user_data:
            self.confirm_password = QLineEdit()
            self.confirm_password.setPlaceholderText("Повторите пароль")
            self.confirm_password.setEchoMode(QLineEdit.EchoMode.Password)
            self.confirm_password.setMinimumHeight(35)
            main_form.addRow("Подтверждение пароля *:", self.confirm_password)

        # ФИО
        self.full_name = QLineEdit()
        self.full_name.setPlaceholderText("Иванов Иван Иванович")
        self.full_name.setMinimumHeight(35)
        main_form.addRow("ФИО *:", self.full_name)

        # Роль
        self.role = QComboBox()
        self.role.addItems(["user", "admin"])
        self.role.setItemText(0, "Пользователь")
        self.role.setItemText(1, "Администратор")
        self.role.setMinimumHeight(35)
        self.role.currentTextChanged.connect(self.on_role_changed)
        main_form.addRow("Роль:", self.role)

        main_group.setLayout(main_form)
        scroll_layout.addWidget(main_group)

        # ========== БЛОК 2: ИНФОРМАЦИЯ О ПОЛЬЗОВАТЕЛЕ ==========
        info_group = QGroupBox("Личная информация")
        info_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 2px solid #2ecc71;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #2ecc71;
            }
        """)

        info_form = QFormLayout()
        info_form.setSpacing(12)
        info_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # Подразделение
        self.department = QComboBox()
        self.department.setEditable(True)
        self.department.setMinimumHeight(35)
        self.load_departments()
        info_form.addRow("Подразделение:", self.department)

        # Должность
        self.position = QComboBox()
        self.position.addItems([
            "Рядовой состав",
            "Сержантский состав",
            "Прапорщики",
            "Младший офицерский состав",
            "Старший офицерский состав",
            "Высший офицерский состав",
            "Гражданский персонал"
        ])
        self.position.setEditable(True)
        self.position.setMinimumHeight(35)
        info_form.addRow("Должность:", self.position)

        # Звание
        self.rank = QComboBox()
        self.rank.addItems([
            "ряд.", "ефр.", "мл. серж.", "серж.", "ст. серж.", "старшина",
            "прапорщик", "ст. прапорщик", "мл. лейт.", "лейтенант", "ст. лейтенант",
            "капитан", "майор", "подполковник", "полковник", "генерал-майор",
            "генерал-лейтенант", "генерал-полковник"
        ])
        self.rank.setEditable(True)
        self.rank.setMinimumHeight(35)
        info_form.addRow("Звание:", self.rank)

        # Чекбокс "Начальник подразделения"
        self.is_head_checkbox = QCheckBox("Начальник подразделения")
        self.is_head_checkbox.setStyleSheet("""
            QCheckBox {
                spacing: 8px;
                font-weight: bold;
                margin-top: 10px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)
        self.is_head_checkbox.setToolTip("Начальник подразделения может:\n"
                                         "• Редактировать план набора\n"
                                         "• Просматривать всех абитуриентов подразделения\n"
                                         "• Управлять группами своего подразделения")
        self.is_head_checkbox.stateChanged.connect(self.on_head_changed)
        info_form.addRow("", self.is_head_checkbox)

        info_group.setLayout(info_form)
        scroll_layout.addWidget(info_group)

        # ========== БЛОК 3: ПРАВА ДОСТУПА ==========
        self.permissions_group = QGroupBox("Права доступа к подразделениям")
        self.permissions_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 2px solid #e67e22;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #e67e22;
            }
        """)

        permissions_layout = QVBoxLayout()

        info_label = QLabel("Выберите подразделения, к которым пользователь будет иметь доступ:")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #7f8c8d; font-size: 11px; margin-bottom: 10px;")
        permissions_layout.addWidget(info_label)

        self.permissions_widget = QWidget()
        self.permissions_grid = QGridLayout(self.permissions_widget)
        self.permissions_grid.setSpacing(10)
        self.permission_checkboxes = []

        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)

        self.select_all_btn = QPushButton("Выбрать все")
        self.select_all_btn.clicked.connect(self.select_all_permissions)
        self.select_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 5px 10px;
                font-size: 11px;
            }
        """)

        self.clear_all_btn = QPushButton("Снять все")
        self.clear_all_btn.clicked.connect(self.clear_all_permissions)
        self.clear_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 5px 10px;
                font-size: 11px;
            }
        """)

        buttons_layout.addWidget(self.select_all_btn)
        buttons_layout.addWidget(self.clear_all_btn)
        buttons_layout.addStretch()

        permissions_layout.addWidget(buttons_widget)
        permissions_layout.addWidget(self.permissions_widget)

        self.permissions_group.setLayout(permissions_layout)
        scroll_layout.addWidget(self.permissions_group)

        scroll_layout.addStretch()
        scroll_area.setWidget(scroll_widget)
        main_layout.addWidget(scroll_area)

        # Информация об обязательных полях
        required_label = QLabel("* - поля, обязательные для заполнения")
        required_label.setStyleSheet("color: #e74c3c; font-size: 11px; margin-top: 5px;")
        main_layout.addWidget(required_label)

        # Кнопки
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )

        # ВАЖНО: правильное подключение сигналов
        ok_button = button_box.button(QDialogButtonBox.StandardButton.Ok)
        ok_button.setText("Сохранить")
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)

        cancel_button = button_box.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)

        # ПОДКЛЮЧАЕМ ПРАВИЛЬНЫЕ СИГНАЛЫ
        button_box.accepted.connect(self.validate_and_accept)  # Сохранение с валидацией
        button_box.rejected.connect(self.reject)  # Отмена - просто закрываем

        main_layout.addWidget(button_box)

        # Заполнение данных при редактировании
        if self.user_data:
            self.load_user_data()

        # Загружаем права доступа
        self.load_permissions()

        # Обновляем видимость блоков в зависимости от роли
        self.on_role_changed(self.role.currentText())

    def load_departments(self):
        """Загрузка подразделений из БД"""
        if self.db:
            departments = self.db.get_departments()
            self.department.clear()
            self.department.addItem("")
            for dept in departments:
                self.department.addItem(dept['name'])

    def load_permissions(self):
        """Загрузка чекбоксов для прав доступа (без дублей)"""
        # Очищаем существующие чекбоксы
        for checkbox in self.permission_checkboxes:
            checkbox.deleteLater()
        self.permission_checkboxes.clear()

        # Очищаем сетку
        while self.permissions_grid.count():
            item = self.permissions_grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Получаем уникальные подразделения из БД
        departments = []
        if self.db:
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT DISTINCT name FROM departments WHERE type != "root" ORDER BY name')
            departments = [row['name'] for row in cursor.fetchall()]

        if not departments:
            info_label = QLabel("Нет доступных подразделений. Создайте их в настройках.")
            info_label.setStyleSheet("color: #e74c3c; font-style: italic;")
            self.permissions_grid.addWidget(info_label, 0, 0)
            return

        # Создаем чекбоксы (без дублей)
        row = 0
        col = 0
        for dept in departments:
            checkbox = QCheckBox(dept)
            self.permission_checkboxes.append(checkbox)
            self.permissions_grid.addWidget(checkbox, row, col)
            col += 1
            if col >= 3:
                col = 0
                row += 1

    def on_role_changed(self, role):
        """Обработка изменения роли"""
        is_admin = (role == "admin")
        self.permissions_group.setVisible(not is_admin)
        self.is_head_checkbox.setVisible(not is_admin)

        if self.is_head_checkbox.isChecked():
            self.permissions_group.setVisible(False)

    def on_head_changed(self, state):
        """Обработка изменения чекбокса начальника"""
        is_head = state == Qt.Checked
        self.permissions_group.setVisible(not is_head)

    def select_all_permissions(self):
        """Выбрать все права"""
        for checkbox in self.permission_checkboxes:
            checkbox.setChecked(True)

    def clear_all_permissions(self):
        """Снять все права"""
        for checkbox in self.permission_checkboxes:
            checkbox.setChecked(False)

    def validate_and_accept(self):
        """Валидация данных перед сохранением"""
        errors = []

        if not self.username.text().strip():
            errors.append("Логин")
            self.username.setStyleSheet("border: 2px solid #e74c3c;")
        else:
            self.username.setStyleSheet("")

        if not self.password.text().strip():
            errors.append("Пароль")
            self.password.setStyleSheet("border: 2px solid #e74c3c;")
        else:
            self.password.setStyleSheet("")

        if not self.user_data:
            if self.password.text() != self.confirm_password.text():
                errors.append("Пароли не совпадают")
                self.password.setStyleSheet("border: 2px solid #e74c3c;")
                self.confirm_password.setStyleSheet("border: 2px solid #e74c3c;")
            else:
                self.confirm_password.setStyleSheet("")

        if not self.full_name.text().strip():
            errors.append("ФИО")
            self.full_name.setStyleSheet("border: 2px solid #e74c3c;")
        else:
            self.full_name.setStyleSheet("")

        if errors:
            error_msg = "Пожалуйста, заполните следующие обязательные поля:\n• " + "\n• ".join(errors)
            QMessageBox.warning(self, "Ошибка валидации", error_msg)
            return

        self.accept()

    def load_user_data(self):
        """Загрузка данных пользователя для редактирования"""
        self.username.setText(self.user_data.get('username', ''))
        self.full_name.setText(self.user_data.get('full_name', ''))

        role = self.user_data.get('role', 'user')
        self.role.setCurrentText(role)

        is_head = self.user_data.get('is_head', False)
        self.is_head_checkbox.setChecked(is_head)

        department_name = self.user_data.get('department_name', '')
        if not department_name and self.user_data.get('department_id'):
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT name FROM departments WHERE id = ?', (self.user_data['department_id'],))
            dept = cursor.fetchone()
            if dept:
                department_name = dept['name']

        if department_name:
            index = self.department.findText(department_name)
            if index >= 0:
                self.department.setCurrentIndex(index)
            else:
                self.department.setEditText(department_name)

        position = self.user_data.get('position', '')
        if position:
            index = self.position.findText(position)
            if index >= 0:
                self.position.setCurrentIndex(index)

        rank = self.user_data.get('rank', '')
        if rank:
            index = self.rank.findText(rank)
            if index >= 0:
                self.rank.setCurrentIndex(index)

        if self.db and role != 'admin' and not is_head:
            permissions = self.db.get_user_department_permissions(self.user_data['id'])
            for checkbox in self.permission_checkboxes:
                for perm in permissions:
                    if checkbox.text() == perm['department_name']:
                        checkbox.setChecked(True)
                        break

    def get_data(self):
        """Получение данных из формы"""
        department_name = self.department.currentText()
        department_id = None
        if self.db and department_name:
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT id FROM departments WHERE name = ?', (department_name,))
            result = cursor.fetchone()
            if result:
                department_id = result['id']

        data = {
            'username': self.username.text().strip(),
            'password': self.password.text().strip(),
            'full_name': self.full_name.text().strip(),
            'role': self.role.currentText(),
            'department_id': department_id,
            'position': self.position.currentText(),
            'rank': self.rank.currentText(),
            'is_head': self.is_head_checkbox.isChecked(),
        }

        if data['role'] != 'admin' and not data['is_head']:
            selected_departments = []
            for checkbox in self.permission_checkboxes:
                if checkbox.isChecked():
                    selected_departments.append(checkbox.text())
            data['permissions'] = selected_departments
        else:
            data['permissions'] = []

        return data


class PermissionDialog(QDialog):
    """Диалог добавления права доступа"""

    def __init__(self, user_id, user_name, parent=None):
        super().__init__(parent)
        self.user_id = user_id
        self.user_name = user_name
        self.setModal(True)
        self.setWindowTitle(f'Добавить права для {user_name}')
        self.setFixedSize(400, 200)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        form_layout = QFormLayout()

        # Поля формы
        self.permission_type = QComboBox()
        self.permission_type.addItems(['all', 'course'])
        self.permission_type.currentTextChanged.connect(self.update_fields)

        self.course = QComboBox()
        self.course.addItems(['1 курс', '2 курс', '3 курс', '4 курс', '5 курс'])

        # self.faculty = QLineEdit()
        # self.faculty.setPlaceholderText('Введите факультет')

        # Скрываем по умолчанию
        self.course.setVisible(False)
        # self.faculty.setVisible(False)

        # Добавление полей в форму
        form_layout.addRow('Тип права:', self.permission_type)
        form_layout.addRow('Курс:', self.course)
        # form_layout.addRow('Факультет:', self.faculty)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        self.setLayout(layout)

    def update_fields(self, permission_type):
        """Обновление видимости полей в зависимости от типа права"""
        self.course.setVisible(permission_type == 'course')

    def get_data(self):
        """Получение данных из формы"""
        permission_type = self.permission_type.currentText()

        data = {
            'user_id': self.user_id,
            'permission_type': permission_type,
            'course': None,
            'faculty': None
        }

        if permission_type == 'course':
            data['course'] = self.course.currentText()
        elif permission_type == 'faculty':
            data['faculty'] = self.faculty.text().strip()

        return data


class ImportWorker(QThread):
    """Поток для импорта данных"""
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str)

    def __init__(self, file_path, password, selected_sheets, mapping, user_id, db):
        super().__init__()
        self.file_path = file_path
        self.password = password
        self.selected_sheets = selected_sheets
        self.mapping = mapping
        self.user_id = user_id
        self.db = db

    def run(self):
        try:
            temp_file_path = None

            # Расшифровка если есть пароль
            if self.password:
                try:
                    import msoffcrypto
                    import tempfile

                    with open(self.file_path, "rb") as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password=self.password)

                        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                        temp_file_path = temp_file.name
                        office_file.decrypt(temp_file)
                        temp_file.close()

                        self.file_path = temp_file_path
                except Exception as e:
                    self.finished.emit(False, f"Ошибка расшифровки: {str(e)}")
                    return

            # Читаем Excel файл
            try:
                xls = pd.ExcelFile(self.file_path, engine='openpyxl')
            except:
                try:
                    xls = pd.ExcelFile(self.file_path, engine='xlrd')
                except:
                    xls = pd.ExcelFile(self.file_path)

            imported_count = 0
            duplicate_count = 0
            error_count = 0
            total_rows = 0

            # Подсчитываем общее количество строк для прогресса
            for sheet_name in self.selected_sheets:
                if sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=0)
                    total_rows += len(df)

            current_row = 0

            for sheet_name in self.selected_sheets:
                if sheet_name not in xls.sheet_names:
                    continue

                self.progress.emit(int((current_row / total_rows) * 100), f"Обработка листа: {sheet_name}")

                df = pd.read_excel(xls, sheet_name=sheet_name, header=0)

                # Определяем курс из названия листа
                course_from_sheet = ''
                for course_num in ['1', '2', '3', '4', '5']:
                    if f'{course_num} курс' in sheet_name.lower():
                        course_from_sheet = f'{course_num} курс'
                        break

                for index, row in df.iterrows():
                    current_row += 1

                    if current_row % 10 == 0:
                        self.progress.emit(int((current_row / total_rows) * 100),
                                           f"Импорт: {current_row}/{total_rows}")

                    # Пропускаем пустые строки
                    if row.isnull().all():
                        continue

                    # Извлекаем данные по маппингу
                    applicant_data = self.extract_data(row, self.mapping, course_from_sheet)

                    if not applicant_data.get('applicant_name'):
                        error_count += 1
                        continue

                    # Проверка на дубликат
                    if self.check_duplicate(applicant_data):
                        duplicate_count += 1
                        continue

                    # Добавляем в БД
                    try:
                        self.db.add_applicant(self.user_id, applicant_data)
                        imported_count += 1
                    except Exception as e:
                        error_count += 1
                        print(f"Ошибка добавления: {e}")

            # Очистка
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except:
                    pass

            result_message = f"""
            Импорт завершен!

            Статистика:
            • Успешно импортировано: {imported_count}
            • Пропущено дубликатов: {duplicate_count}
            • Ошибок: {error_count}
            • Обработано листов: {len(self.selected_sheets)}
            """

            self.finished.emit(True, result_message)

        except Exception as e:
            self.finished.emit(False, f"Ошибка импорта: {str(e)}")

    def extract_data(self, row, mapping, course_from_sheet):
        """Извлечение данных по маппингу"""
        data = {
            'applicant_name': '',
            'region': '',
            'city': '',
            'category': '',
            'phone': '',
            'education': '',
            'status': 'поступает',
            'document_status': '',
            'agitator_department': '',
            'agitator_name': '',
            'agitator_course': '',
            'agitator_group': '',
            'agitator_rank': '',
            'agitator_is_cadet': False,
            'notes': ''
        }

        # Маппинг полей - ВАЖНО: правильно сопоставляем!
        for field, column in mapping.items():
            if column and column in row and pd.notna(row[column]):
                value = str(row[column]).strip()

                if field == 'applicant_name':
                    data['applicant_name'] = value
                elif field == 'region':
                    data['region'] = value
                elif field == 'city':
                    data['city'] = value
                elif field == 'category':
                    if value.lower() in ['м', 'м.', 'муж', 'мужчина', 'male']:
                        data['category'] = 'м'
                    elif value.lower() in ['ж', 'ж.', 'жен', 'женщина', 'female']:
                        data['category'] = 'ж'
                    elif value.lower() in ['всл', 'военнослужащий', 'военнослужащие', 'воен']:
                        data['category'] = 'всл'
                    else:
                        data['category'] = value[:2].lower()
                elif field == 'phone':
                    data['phone'] = self.normalize_phone(value)
                elif field == 'education':
                    data['education'] = value
                elif field == 'status':
                    if 'поступает' in value.lower() or 'поступают' in value.lower():
                        data['status'] = 'поступает'
                    else:
                        data['status'] = 'отказывается'
                elif field == 'document_status':
                    data['document_status'] = value
                elif field == 'agitator_department':
                    data['agitator_department'] = value
                elif field == 'agitator_name':
                    data['agitator_name'] = value
                elif field == 'agitator_course':
                    data['agitator_course'] = value
                elif field == 'agitator_group':
                    data['agitator_group'] = value
                elif field == 'agitator_rank':
                    data['agitator_rank'] = value
                elif field == 'notes':
                    data['notes'] = value
        # Если курс не указан, берем из названия листа
        if not data['agitator_course'] and course_from_sheet:
            data['agitator_course'] = course_from_sheet

        # Определяем тип агитатора
        if data['agitator_group'] or (data['agitator_course'] and not data['agitator_rank']):
            data['agitator_is_cadet'] = True
        return data

    @staticmethod
    def normalize_phone(phone):
        """Нормализация номера телефона"""
        if not phone:
            return ""
        digits = ''.join(filter(str.isdigit, phone))
        if len(digits) < 10:
            return phone
        if digits.startswith('8') and len(digits) == 11:
            return '7' + digits[1:]
        elif len(digits) == 10:
            return '7' + digits
        elif digits.startswith('7') and len(digits) == 11:
            return digits
        return digits

    def check_duplicate(self, applicant_data):
        """Проверка на дубликат"""
        cursor = self.db.conn.cursor()
        cursor.execute('''
            SELECT COUNT(*) FROM applicants 
            WHERE applicant_name = ? AND phone = ?
        ''', (applicant_data['applicant_name'], applicant_data['phone']))
        count = cursor.fetchone()[0]
        return count > 0


class ImportDialog(QDialog):
    """Диалог для импорта данных с маппингом колонок"""

    def __init__(self, db, user_id, parent=None):
        super().__init__(parent)
        self.db = db
        self.user_id = user_id
        self.file_path = None
        self.available_columns = []
        self.mapping = {}
        self.setModal(True)
        self.setWindowTitle('Импорт данных из Excel')
        self.setMinimumSize(800, 700)
        self.setMaximumSize(1000, 800)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # Заголовок
        title = QLabel("Импорт данных из Excel")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_font = title.font()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)

        # Создаем скролл область для всего содержимого
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("QScrollArea { border: none; background-color: transparent; }")

        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(20)

        # 1. Выбор файла
        file_group = QGroupBox("1. Выбор файла")
        file_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #3498db;
            }
        """)
        file_layout = QHBoxLayout()

        self.file_label = QLabel("Файл не выбран")
        self.file_label.setStyleSheet("color: #e74c3c; padding: 5px;")

        self.select_file_btn = QPushButton("Выбрать файл")
        self.select_file_btn.clicked.connect(self.select_file)
        self.select_file_btn.setMinimumHeight(35)

        file_layout.addWidget(self.select_file_btn)
        file_layout.addWidget(self.file_label, 1)
        file_group.setLayout(file_layout)
        scroll_layout.addWidget(file_group)

        # 2. Пароль
        password_group = QGroupBox("2. Пароль (если требуется)")
        password_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #e67e22;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #e67e22;
            }
        """)
        password_layout = QHBoxLayout()

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Введите пароль для защищенного файла")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setMinimumHeight(35)

        password_layout.addWidget(self.password_input)
        password_group.setLayout(password_layout)
        scroll_layout.addWidget(password_group)

        # 3. Выбор листов
        sheets_group = QGroupBox("3. Выбор листов для импорта")
        sheets_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #2ecc71;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #2ecc71;
            }
        """)
        sheets_layout = QVBoxLayout()

        self.load_sheets_btn = QPushButton("Загрузить листы")
        self.load_sheets_btn.clicked.connect(self.load_sheets)
        self.load_sheets_btn.setEnabled(False)
        self.load_sheets_btn.setMinimumHeight(35)
        sheets_layout.addWidget(self.load_sheets_btn)

        self.sheets_widget = QWidget()
        self.sheets_layout = QVBoxLayout(self.sheets_widget)
        self.sheets_widget.setVisible(False)
        sheets_layout.addWidget(self.sheets_widget)

        sheets_group.setLayout(sheets_layout)
        scroll_layout.addWidget(sheets_group)

        # 4. Сопоставление колонок
        mapping_group = QGroupBox("4. Сопоставление колонок")
        mapping_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #9b59b6;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                color: #9b59b6;
            }
        """)

        # Создаем скролл для маппинга
        mapping_scroll = QScrollArea()
        mapping_scroll.setWidgetResizable(True)
        mapping_scroll.setMaximumHeight(400)
        mapping_scroll.setStyleSheet("QScrollArea { border: 1px solid #ddd; border-radius: 5px; }")

        self.mapping_widget = QWidget()
        self.mapping_layout = QFormLayout(self.mapping_widget)
        self.mapping_layout.setSpacing(10)
        self.mapping_widget.setVisible(False)

        mapping_scroll.setWidget(self.mapping_widget)

        mapping_layout_main = QVBoxLayout()
        mapping_layout_main.addWidget(mapping_scroll)
        mapping_group.setLayout(mapping_layout_main)
        scroll_layout.addWidget(mapping_group)

        scroll_layout.addStretch()
        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # Кнопки
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )

        self.ok_button = button_box.button(QDialogButtonBox.StandardButton.Ok)
        self.ok_button.setText("Начать импорт")
        self.ok_button.setEnabled(False)
        self.ok_button.setMinimumHeight(40)
        self.ok_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)

        cancel_button = button_box.button(QDialogButtonBox.StandardButton.Cancel)
        cancel_button.setText("Отмена")
        cancel_button.setMinimumHeight(40)

        button_box.accepted.connect(self.start_import)
        button_box.rejected.connect(self.reject)

        layout.addWidget(button_box)

    def select_file(self):
        """Выбор Excel файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 'Выберите Excel файл', '',
            'Excel Files (*.xlsx *.xls *.xlsm);;All Files (*)'
        )

        if file_path:
            self.file_path = file_path
            self.file_label.setText(f"{os.path.basename(file_path)}")
            self.file_label.setStyleSheet("color: #2ecc71; padding: 5px;")
            self.load_sheets_btn.setEnabled(True)
            self.reset_sheets()

    def reset_sheets(self):
        """Сброс выбора листов"""
        # Очищаем чекбоксы
        while self.sheets_layout.count():
            item = self.sheets_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        self.sheets_widget.setVisible(False)
        self.mapping_widget.setVisible(False)
        self.ok_button.setEnabled(False)
        self.available_columns = []
        self.mapping = {}

    def load_sheets(self):
        """Загрузка списка листов"""
        if not self.file_path:
            return

        try:
            # Определяем движок
            engine = 'openpyxl' if self.file_path.endswith('.xlsx') else 'xlrd'

            # Пробуем прочитать с паролем
            temp_file_path = None
            file_path = self.file_path

            if self.password_input.text().strip():
                try:
                    import msoffcrypto
                    import tempfile

                    with open(self.file_path, "rb") as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password=self.password_input.text().strip())

                        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                        temp_file_path = temp_file.name
                        office_file.decrypt(temp_file)
                        temp_file.close()

                        file_path = temp_file_path
                except Exception as e:
                    QMessageBox.warning(self, "Ошибка", f"Неверный пароль: {str(e)}")
                    return

            xls = pd.ExcelFile(file_path, engine=engine)

            # Очищаем старые чекбоксы
            while self.sheets_layout.count():
                item = self.sheets_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

            # Создаем чекбоксы для каждого листа
            self.sheet_checkboxes = []
            sheets_container = QWidget()
            sheets_container_layout = QGridLayout(sheets_container)

            row = 0
            col = 0
            for sheet in xls.sheet_names:
                checkbox = QCheckBox(sheet)
                if 'курс' in sheet.lower():
                    checkbox.setChecked(True)
                self.sheet_checkboxes.append(checkbox)
                sheets_container_layout.addWidget(checkbox, row, col)
                col += 1
                if col >= 3:
                    col = 0
                    row += 1

            self.sheets_layout.addWidget(sheets_container)

            # Добавляем кнопки выбора
            buttons_widget = QWidget()
            buttons_layout = QHBoxLayout(buttons_widget)

            select_all_btn = QPushButton("Выбрать все")
            select_all_btn.clicked.connect(self.select_all_sheets)
            clear_all_btn = QPushButton("Снять все")
            clear_all_btn.clicked.connect(self.clear_all_sheets)

            buttons_layout.addWidget(select_all_btn)
            buttons_layout.addWidget(clear_all_btn)
            buttons_layout.addStretch()

            self.sheets_layout.addWidget(buttons_widget)
            self.sheets_widget.setVisible(True)

            # Сохраняем колонки для маппинга
            first_sheet = xls.sheet_names[0]
            df = pd.read_excel(file_path, sheet_name=first_sheet, header=0)
            self.available_columns = list(df.columns)

            # Создаем маппинг
            self.create_mapping_ui()

            # Очищаем временный файл
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except:
                    pass

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить листы: {str(e)}")

    def select_all_sheets(self):
        """Выбрать все листы"""
        for checkbox in self.sheet_checkboxes:
            checkbox.setChecked(True)

    def clear_all_sheets(self):
        """Снять все выделения"""
        for checkbox in self.sheet_checkboxes:
            checkbox.setChecked(False)

    def create_mapping_ui(self):
        """Создание интерфейса для маппинга колонок"""
        # Очищаем старый маппинг
        while self.mapping_layout.count():
            item = self.mapping_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Поля для маппинга - правильный порядок!
        fields = [
            ('applicant_name', 'ФИО абитуриента *', True),
            ('region', 'Субъект РФ', False),
            ('city', 'Населенный пункт', False),
            ('category', 'Категория (м/ж/всл)', False),
            ('phone', 'Телефон', False),
            ('education', 'Образование', False),
            ('status', 'Статус (поступает/отказывается)', False),
            ('document_status', 'Документы', False),
            ('agitator_department', 'Подразделение агитатора', False),
            ('agitator_name', 'ФИО агитатора *', True),
            ('agitator_course', 'Курс агитатора', False),
            ('agitator_group', 'Группа агитатора', False),
            ('agitator_rank', 'Звание агитатора', False),
            ('notes', 'Примечания', False)
        ]

        # Показываем доступные колонки для отладки
        for field, label, required in fields:
            # Создаем виджет для строки
            row_widget = QFrame()
            row_widget.setFrameStyle(QFrame.Shape.StyledPanel)
            row_widget.setStyleSheet("QFrame { background-color: #f8f9fa; border-radius: 5px; }")

            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(10, 5, 10, 5)

            # Метка
            field_label = QLabel(label)
            field_label.setMinimumWidth(200)
            if required:
                field_label.setStyleSheet("color: #e74c3c; font-weight: bold;")

            # Комбобокс с колонками
            combo = QComboBox()
            combo.addItem("-- Не выбрано --")
            combo.addItems(self.available_columns)
            combo.setMinimumWidth(300)
            combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            # Автоматическое определение колонки
            auto_match = self.auto_match_column(field)
            if auto_match and auto_match in self.available_columns:
                index = combo.findText(auto_match)
                if index >= 0:
                    combo.setCurrentIndex(index)
            row_layout.addWidget(field_label)
            row_layout.addWidget(combo, 1)

            self.mapping_layout.addRow(row_widget)

            # Сохраняем ссылку на комбобокс
            setattr(self, f"combo_{field}", combo)

        # Добавляем пояснение
        info_label = QLabel("* Обязательные поля для сопоставления")
        info_label.setStyleSheet("color: #7f8c8d; font-style: italic; margin-top: 10px;")
        self.mapping_layout.addRow(info_label)

        # Добавляем кнопку для проверки маппинга
        test_btn = QPushButton("Проверить маппинг")
        test_btn.clicked.connect(self.test_mapping)
        self.mapping_layout.addRow(test_btn)

        self.mapping_widget.setVisible(True)
        self.ok_button.setEnabled(True)

    def test_mapping(self):
        """Проверка маппинга - показывает первые 5 строк данных"""
        if not hasattr(self, 'sheet_checkboxes') or not self.sheet_checkboxes:
            QMessageBox.warning(self, "Ошибка", "Сначала загрузите листы!")
            return

        # Берем первый выбранный лист
        selected = [cb for cb in self.sheet_checkboxes if cb.isChecked()]
        if not selected:
            QMessageBox.warning(self, "Ошибка", "Выберите хотя бы один лист!")
            return

        sheet_name = selected[0].text()

        try:
            # Читаем файл
            file_path = self.file_path
            temp_file_path = None

            if self.password_input.text().strip():
                try:
                    import msoffcrypto
                    import tempfile

                    with open(self.file_path, "rb") as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password=self.password_input.text().strip())

                        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                        temp_file_path = temp_file.name
                        office_file.decrypt(temp_file)
                        temp_file.close()

                        file_path = temp_file_path
                except:
                    pass

            df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

            # Показываем первые 5 строк
            preview_text = f"Первые 5 строк из листа '{sheet_name}':\n\n"
            preview_text += df.head(5).to_string()

            # Также показываем маппинг
            mapping = self.get_mapping()
            preview_text += f"\n\nТекущий маппинг:\n"
            for field, column in mapping.items():
                if column:
                    preview_text += f"  {field} -> {column}\n"
                else:
                    preview_text += f"  {field} -> (не выбрано)\n"

            QMessageBox.information(self, "Проверка данных", preview_text)

            # Очищаем временный файл
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                except:
                    pass

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось прочитать файл: {str(e)}")

    def auto_match_column(self, field):
        """Автоматическое определение колонки по названию"""
        matches = {
            'applicant_name': ['абитуриент', 'фио абитуриента', 'фио', 'ф.и.о.', 'фио студента', 'фио абитуриента'],
            'region': ['субъект', 'регион', 'область', 'край', 'республика', 'субъект рф'],
            'city': ['город', 'населенный пункт', 'населённый пункт', 'город/село', 'населенный', 'город'],
            'category': ['категория', 'пол', 'кат', 'категория абитуриента', 'категория'],
            'phone': ['телефон', 'тел', 'номер телефона', 'контактный телефон', 'мобильный', 'телефон'],
            'education': ['образование', 'уровень образования', 'школа', 'вуз', 'училище', 'образование'],
            'status': ['статус', 'поступление', 'статус поступления', 'решение', 'статус'],
            'document_status': ['документы', 'статус документов', 'личное дело', 'документ', 'документы'],
            'agitator_department': ['подразделение агитатора', 'подразделение', 'кафедра', 'факультет', 'отделение',
                                    'подр'],
            'agitator_name': ['агитатор', 'фио агитатора', 'кто пригласил', 'пригласил', 'агитатор фио',
                              'фио агитатора'],
            'agitator_course': ['курс агитатора', 'курс', 'год обучения', 'курс'],
            'agitator_group': ['группа агитатора', 'группа', 'учебная группа', 'номер группы', 'группа'],
            'agitator_rank': ['звание агитатора', 'звание', 'воинское звание', 'в/з', 'ранг', 'звание'],
            'notes': ['примечание', 'комментарий', 'заметки', 'прим', 'примечания']
        }

        field_matches = matches.get(field, [])
        for col in self.available_columns:
            col_lower = str(col).lower().strip()
            for match in field_matches:
                if match.lower() in col_lower:
                    return col
        return None

    def get_selected_sheets(self):
        """Получить выбранные листы"""
        if not hasattr(self, 'sheet_checkboxes'):
            return []
        return [cb.text() for cb in self.sheet_checkboxes if cb.isChecked()]

    def get_mapping(self):
        """Получить маппинг колонок"""
        mapping = {}
        fields = ['applicant_name', 'region', 'city', 'category', 'phone',
                  'education', 'status', 'document_status', 'agitator_department',
                  'agitator_name', 'agitator_course', 'agitator_group',
                  'agitator_rank', 'notes']

        for field in fields:
            combo = getattr(self, f"combo_{field}", None)
            if combo and combo.currentIndex() > 0:
                mapping[field] = combo.currentText()
            else:
                mapping[field] = None

        return mapping

    def start_import(self):
        """Запуск импорта"""
        selected_sheets = self.get_selected_sheets()
        if not selected_sheets:
            QMessageBox.warning(self, "Ошибка", "Выберите хотя бы один лист для импорта!")
            return

        mapping = self.get_mapping()

        # Проверяем обязательные поля
        if not mapping.get('applicant_name'):
            QMessageBox.warning(self, "Ошибка", "Необходимо сопоставить колонку 'ФИО абитуриента'!")
            return

        if not mapping.get('agitator_name'):
            QMessageBox.warning(self, "Ошибка", "Необходимо сопоставить колонку 'ФИО агитатора'!")
            return

        # Запускаем импорт в отдельном потоке
        self.worker = ImportWorker(
            self.file_path,
            self.password_input.text().strip(),
            selected_sheets,
            mapping,
            self.user_id,
            self.db
        )

        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.import_finished)

        # Показываем прогресс
        self.progress_dialog = QProgressDialog("Подготовка к импорту...", "Отмена", 0, 100, self)
        self.progress_dialog.setWindowTitle("Импорт данных")
        self.progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        self.progress_dialog.canceled.connect(self.worker.terminate)
        self.progress_dialog.show()

        self.worker.start()

    def update_progress(self, value, message):
        """Обновление прогресса"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.setValue(value)
            self.progress_dialog.setLabelText(message)

    def import_finished(self, success, message):
        """Завершение импорта"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()

        if success:
            QMessageBox.information(self, "Успех", message)
            self.accept()
        else:
            QMessageBox.critical(self, "Ошибка", message)


class DepartmentDialog(QDialog):
    """Диалог добавления/редактирования подразделения"""

    def __init__(self, dept_data=None, db=None, parent=None):
        super().__init__(parent)
        self.dept_data = dept_data
        self.db = db
        self.setModal(True)

        if dept_data:
            self.setWindowTitle('Редактировать подразделение')
        else:
            self.setWindowTitle('Добавить подразделение')

        self.setMinimumSize(450, 400)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        form_layout = QFormLayout()
        form_layout.setSpacing(15)

        # Название подразделения
        self.name = QLineEdit()
        self.name.setPlaceholderText("Например: Факультет 1, Кафедра 1, Группа 101")
        if self.dept_data:
            self.name.setText(self.dept_data.get('name', ''))
        form_layout.addRow("Название *:", self.name)

        # Тип подразделения
        self.dept_type = QComboBox()
        self.dept_type.addItems(["faculty", "department", "group"])
        self.dept_type.setItemText(0, "Факультет")
        self.dept_type.setItemText(1, "Кафедра")
        self.dept_type.setItemText(2, "Группа")
        if self.dept_data:
            type_index = {"faculty": 0, "department": 1, "group": 2}.get(self.dept_data.get('type', 'department'), 1)
            self.dept_type.setCurrentIndex(type_index)
        self.dept_type.currentTextChanged.connect(self.on_type_changed)
        form_layout.addRow("Тип:", self.dept_type)

        # Родительское подразделение (для групп)
        self.parent_dept = QComboBox()
        self.parent_dept.addItem("Нет (корневое)")
        self.load_parent_departments()
        if self.dept_data and self.dept_data.get('parent_id'):
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT name FROM departments WHERE id = ?', (self.dept_data['parent_id'],))
            parent = cursor.fetchone()
            if parent:
                index = self.parent_dept.findText(parent['name'])
                if index >= 0:
                    self.parent_dept.setCurrentIndex(index)
        form_layout.addRow("Родительское подразделение:", self.parent_dept)

        # Начальник подразделения
        self.head_user = QComboBox()
        self.head_user.addItem("Не назначен")
        self.load_users()
        if self.dept_data and self.dept_data.get('head_user_id'):
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT full_name FROM users WHERE id = ?', (self.dept_data['head_user_id'],))
            head = cursor.fetchone()
            if head:
                index = self.head_user.findText(head['full_name'])
                if index >= 0:
                    self.head_user.setCurrentIndex(index)
        form_layout.addRow("Начальник подразделения:", self.head_user)

        layout.addLayout(form_layout)

        # Кнопки
        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        self.setLayout(layout)
        self.on_type_changed()

    def on_type_changed(self):
        """При изменении типа показываем/скрываем родительское подразделение"""
        is_group = self.dept_type.currentText() == "group"
        self.parent_dept.setVisible(is_group)
        # Меняем метку
        if is_group:
            self.parent_dept.setToolTip("Выберите факультет или кафедру, к которому относится группа")
        else:
            self.parent_dept.setToolTip("")

    def load_parent_departments(self):
        """Загрузка родительских подразделений (только факультеты и кафедры)"""
        if self.db:
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT name FROM departments 
                WHERE type IN ('faculty', 'department') AND type != 'root'
                ORDER BY 
                    CASE type
                        WHEN 'faculty' THEN 1
                        WHEN 'department' THEN 2
                        ELSE 3
                    END,
                    name
            ''')
            for row in cursor.fetchall():
                self.parent_dept.addItem(row['name'])

    def load_users(self):
        """Загрузка пользователей для назначения начальником"""
        if self.db:
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT id, full_name, role 
                FROM users 
                WHERE role != 'admin' OR role = 'admin'
                ORDER BY full_name
            ''')
            for row in cursor.fetchall():
                role_mark = " (Админ)" if row['role'] == 'admin' else ""
                self.head_user.addItem(f"{row['full_name']}{role_mark}", row['id'])

    def get_data(self):
        """Получение данных из формы"""
        data = {
            'name': self.name.text().strip(),
            'type': self.dept_type.currentText(),
            'parent_id': None,
            'head_user_id': None
        }

        # Проверка обязательного поля
        if not data['name']:
            QMessageBox.warning(self, "Ошибка", "Введите название подразделения!")
            return None

        # Родительское подразделение
        parent_name = self.parent_dept.currentText()
        if parent_name and parent_name != "Нет (корневое)" and self.db:
            cursor = self.db.conn.cursor()
            cursor.execute('SELECT id FROM departments WHERE name = ?', (parent_name,))
            parent = cursor.fetchone()
            if parent:
                data['parent_id'] = parent['id']

        # Начальник
        head_data = self.head_user.currentData()
        if head_data:
            data['head_user_id'] = head_data

        return data


class MainWindow(QMainWindow):
    """Главное окно приложения"""

    # Сигнал для выхода из системы
    logout_requested = pyqtSignal()

    def __init__(self, user_data):
        super().__init__()
        icon_path = get_icon_path("icon.ico")
        self.advanced_filters = {}
        if icon_path and os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.user_data = user_data
        self.db = Database()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f'Агитация - {self.user_data["full_name"]}')
        self.setGeometry(100, 100, 1200, 700)

        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Панель инструментов
        toolbar = QToolBar()
        self.addToolBar(toolbar)

        # Действия
        add_action = QAction(QIcon(resource_path("icons/add_document.png")), 'Добавить', self)
        add_action.triggered.connect(self.add_applicant)
        toolbar.addAction(add_action)

        edit_action = QAction(QIcon(resource_path("icons/pencel.png")), 'Редактировать', self)
        edit_action.triggered.connect(self.edit_applicant)
        toolbar.addAction(edit_action)

        delete_action = QAction(QIcon(resource_path("icons/delete_document.png")), 'Удалить', self)
        delete_action.triggered.connect(self.delete_applicant)
        toolbar.addAction(delete_action)

        toolbar.addSeparator()

        # # Кнопка выхода
        # logout_action = QAction(QIcon(resource_path("icons/logout.png")), 'Выход', self)
        # logout_action.triggered.connect(self.logout)
        # toolbar.addAction(logout_action)

        import_action = QAction(QIcon(resource_path("icons/import.png")), 'Импорт из Excel', self)
        import_action.triggered.connect(self.import_from_excel)
        toolbar.addAction(import_action)

        export_action = QAction(QIcon(resource_path("icons/export.png")), 'Экспорт', self)
        export_action.triggered.connect(self.export_data)
        toolbar.addAction(export_action)

        # Растягивающийся разделитель (толкает все кнопки справа налево)
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        toolbar.addWidget(spacer)

        # Кнопка выхода (будет справа)
        logout_action = QAction(QIcon(resource_path("icons/logout.png")), 'Выход', self)
        logout_action.triggered.connect(self.logout)
        toolbar.addAction(logout_action)

        # Вкладки
        self.tab_widget = QTabWidget()

        # Вкладка с данными
        self.data_tab = QWidget()
        self.init_data_tab()
        self.tab_widget.addTab(self.data_tab, QIcon(resource_path("icons/information.png")), 'Данные абитуриентов')

        # Вкладка со статистикой
        self.stats_tab = StatisticsWidget(
            self.user_data['id'],
            self.user_data['role'],
            self.db
        )
        self.tab_widget.addTab(self.stats_tab, QIcon(resource_path("icons/stata.png")), 'Статистика')

        # Вкладка настроек (только для админа)
        if self.user_data['role'] == 'admin':
            self.settings_tab = QWidget()
            self.init_settings_tab()
            self.tab_widget.addTab(self.settings_tab, QIcon(resource_path("icons/settings.png")), 'Настройки (админ)')

        main_layout.addWidget(self.tab_widget)

        # Статус бар
        status_bar = QStatusBar()
        self.setStatusBar(status_bar)
        status_bar.showMessage(f'Пользователь: {self.user_data["full_name"]} | Роль: {self.user_data["role"]}')

        # Обновление данных
        self.refresh_data()

    def open_advanced_search(self):
        """Открытие диалога расширенного поиска"""
        dialog = AdvancedSearchDialog(self.db, self)

        # Если есть сохраненные фильтры, загружаем их (опционально)
        if self.advanced_filters:
            # Здесь можно восстановить предыдущие фильтры
            pass

        if dialog.exec():
            self.advanced_filters = dialog.get_filters()
            self.refresh_data()

            # Показываем информацию о примененных фильтрах
            if self.advanced_filters:
                filter_count = len(self.advanced_filters)
                self.statusBar().showMessage(f"Применено расширенных фильтров: {filter_count}", 3000)
                # Подсвечиваем кнопку
                self.advanced_search_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #8e44ad;
                        color: white;
                        border: 2px solid #f39c12;
                        border-radius: 5px;
                        padding: 8px 16px;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #7d3c98;
                    }
                """)
            else:
                self.reset_filters_style()

    def reset_all_filters(self):
        """Сброс всех фильтров"""
        # Сбрасываем обычный поиск
        self.search_input.clear()

        # Сбрасываем расширенные фильтры
        self.advanced_filters = {}

        # Сбрасываем фильтр по курсу (если есть)
        if hasattr(self, 'course_filter'):
            self.course_filter.setCurrentIndex(0)

        # Обновляем таблицу
        self.refresh_data()

        # Сбрасываем стиль кнопки
        self.reset_filters_style()

        self.statusBar().showMessage("Все фильтры сброшены", 2000)

    def reset_filters_style(self):
        """Сброс стиля кнопки расширенного поиска"""
        self.advanced_search_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)

    def logout(self):
        """Выход из системы"""
        reply = QMessageBox.question(
            self, 'Подтверждение выхода',
            'Вы уверены, что хотите выйти из системы?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            self.db.close()
            self.logout_requested.emit()  # Отправляем сигнал
            self.close()

    def init_data_tab(self):
        """Инициализация вкладки с данными (обновленная версия с панелью выделения)"""
        layout = QVBoxLayout()

        # Панель фильтров
        filter_widget = QWidget()
        filter_layout = QHBoxLayout()

        # Фильтр по курсу (только для админа)
        # if self.user_data['role'] == 'admin':
        #     filter_label = QLabel('Фильтр по курсу:')
        #     self.course_filter = QComboBox()
        #     self.course_filter.addItems(['Все курсы', '1 курс', '2 курс', '3 курс', '4 курс', '5 курс'])
        #     self.course_filter.currentTextChanged.connect(self.refresh_data)
        #
        #     filter_layout.addWidget(filter_label)
        #     filter_layout.addWidget(self.course_filter)

        filter_layout.addStretch()

        # Обычный поиск
        search_label = QLabel('Поиск:')
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('Введите текст для поиска...')
        self.search_input.textChanged.connect(self.refresh_data)
        self.search_input.setMinimumWidth(250)

        # Кнопка расширенного поиска
        self.advanced_search_btn = QPushButton("Расширенный поиск")
        self.advanced_search_btn.setStyleSheet("""
                QPushButton {
                    background-color: #9b59b6;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #8e44ad;
                }
            """)
        self.advanced_search_btn.clicked.connect(self.open_advanced_search)

        # Кнопка сброса фильтров
        self.reset_filters_btn = QPushButton("Сбросить фильтры")
        self.reset_filters_btn.setStyleSheet("""
                QPushButton {
                    background-color: #e67e22;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px 16px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #d35400;
                }
            """)
        self.reset_filters_btn.clicked.connect(self.reset_all_filters)

        filter_layout.addWidget(search_label)
        filter_layout.addWidget(self.search_input)
        filter_layout.addWidget(self.advanced_search_btn)
        filter_layout.addWidget(self.reset_filters_btn)

        filter_widget.setLayout(filter_layout)
        layout.addWidget(filter_widget)

        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "ID",  # скрытая
            "ФИО абитуриента",
            "Субъект РФ",
            "Населенный пункт",
            "Категория",
            "Телефон",
            "Образование",
            "Статус",
            "Документы",
            "Подразделение агитатора",
            "ФИО агитатора",
            "Дата добавления"
        ])

        # Настройка таблицы
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)

        # Скрываем колонку ID
        self.table.setColumnHidden(0, True)

        # Устанавливаем ширину колонок
        self.table.setColumnWidth(1, 220)  # ФИО абитуриента (чуть меньше)
        self.table.setColumnWidth(2, 130)  # Субъект РФ
        self.table.setColumnWidth(3, 120)  # Населенный пункт
        self.table.setColumnWidth(4, 80)  # Категория
        self.table.setColumnWidth(5, 110)  # Телефон
        self.table.setColumnWidth(6, 90)  # Образование
        self.table.setColumnWidth(7, 90)  # Статус
        self.table.setColumnWidth(8, 90)  # Документы
        self.table.setColumnWidth(9, 130)  # Подразделение агитатора
        self.table.setColumnWidth(10, 180)  # ФИО агитатора
        self.table.setColumnWidth(11, 120)  # Дата изменения

        layout.addWidget(self.table)

        # Панель статуса выделения
        self.selection_status_widget = QWidget()
        self.selection_status_widget.setStyleSheet("""
            QWidget {
                background-color: #e8f4fc;
                border-radius: 6px;
                margin-top: 5px;
            }
        """)
        selection_layout = QHBoxLayout(self.selection_status_widget)
        selection_layout.setContentsMargins(15, 10, 15, 10)

        self.selection_info_label = QLabel("Выделено: 0 записей")
        self.selection_info_label.setStyleSheet("font-weight: bold; color: #2c3e50;")

        self.show_stats_btn = QPushButton("Показать статистику по выделенным")
        self.show_stats_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        self.show_stats_btn.clicked.connect(self.show_selection_stats)
        self.show_stats_btn.setEnabled(False)

        self.clear_selection_btn = QPushButton("Снять выделение")
        self.clear_selection_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 6px 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        self.clear_selection_btn.clicked.connect(self.clear_selection)

        selection_layout.addWidget(self.selection_info_label)
        selection_layout.addStretch()
        selection_layout.addWidget(self.show_stats_btn)
        selection_layout.addWidget(self.clear_selection_btn)

        layout.addWidget(self.selection_status_widget)

        self.data_tab.setLayout(layout)

    def on_selection_changed(self):
        """Обработка изменения выделения в таблице"""
        selected_rows = set()
        for item in self.table.selectedItems():
            selected_rows.add(item.row())

        count = len(selected_rows)
        self.selection_info_label.setText(f"Выделено: {count} записей")
        self.show_stats_btn.setEnabled(count > 0)

        # Обновляем статусную строку
        if count > 0:
            self.statusBar().showMessage(f"Выделено {count} записей. Нажмите 'Показать статистику' для анализа.")
        else:
            self.statusBar().showMessage(
                f"Пользователь: {self.user_data['full_name']} | Роль: {self.user_data['role']}")

    def clear_selection(self):
        """Снять все выделения"""
        self.table.clearSelection()
        self.on_selection_changed()

    def show_selection_stats(self):
        """Показать статистику по выделенным строкам"""
        selected_rows = set()
        for item in self.table.selectedItems():
            selected_rows.add(item.row())

        if not selected_rows:
            QMessageBox.warning(self, "Внимание", "Нет выделенных записей!")
            return

        # Собираем данные выделенных строк
        selected_data = []
        for row in selected_rows:
            # Получаем ID записи (первая колонка)
            id_item = self.table.item(row, 0)
            if not id_item or not id_item.text().strip():
                continue

            try:
                applicant_id = int(id_item.text())
            except ValueError:
                continue

            # Получаем полные данные из БД по ID
            cursor = self.db.conn.cursor()
            cursor.execute('''
                SELECT 
                    applicant_name,
                    region,
                    city,
                    category,
                    phone,
                    education,
                    status,
                    document_status,
                    agitator_department,
                    agitator_name,
                    agitator_course,
                    agitator_group,
                    agitator_rank,
                    agitator_is_cadet,
                    created_at,
                    updated_at
                FROM applicants 
                WHERE id = ?
            ''', (applicant_id,))

            result = cursor.fetchone()
            if result:
                row_data = dict(result)
                row_data['id'] = applicant_id
                selected_data.append(row_data)

        if not selected_data:
            QMessageBox.warning(self, "Внимание", "Не удалось получить данные для выделенных записей!")
            return

        # Показываем диалог со статистикой
        stats_dialog = SelectionStatsDialog(selected_data, self)
        stats_dialog.show()

    def refresh_data(self):
        """Обновление данных в таблице"""
        # Сохраняем выделенные ID
        selected_ids = set()
        for item in self.table.selectedItems():
            row = item.row()
            id_item = self.table.item(row, 0)
            if id_item:
                selected_ids.add(int(id_item.text()))

        # Получение данных из БД с учетом расширенных фильтров
        if self.user_data['role'] == 'admin':
            course = self.course_filter.currentText() if hasattr(self, 'course_filter') else None
            applicants = self.db.get_applicants(
                self.user_data['id'],
                'admin',
                course if course != 'Все курсы' else None,
                self.advanced_filters if self.advanced_filters else None
            )
        else:
            applicants = self.db.get_applicants(
                self.user_data['id'],
                self.user_data['role'],
                None,
                self.advanced_filters if self.advanced_filters else None
            )

        # Обычный поиск (дополнительно к расширенному)
        search_text = self.search_input.text().lower().strip()
        if search_text and not self.advanced_filters:
            # Если нет расширенных фильтров, применяем обычный поиск
            filtered_applicants = []
            for applicant in applicants:
                applicant_dict = dict(applicant)
                text_fields = [
                    str(applicant_dict.get('applicant_name', '')),
                    str(applicant_dict.get('region', '')),
                    str(applicant_dict.get('city', '')),
                    str(applicant_dict.get('phone', '')),
                    str(applicant_dict.get('agitator_name', '')),
                    str(applicant_dict.get('agitator_department', '')),
                ]
                if any(search_text in field.lower() for field in text_fields):
                    filtered_applicants.append(applicant)
            applicants = filtered_applicants
        elif search_text and self.advanced_filters:
            # Если есть расширенные фильтры, обычный поиск работает поверх них
            filtered_applicants = []
            for applicant in applicants:
                applicant_dict = dict(applicant)
                text_fields = [
                    str(applicant_dict.get('applicant_name', '')),
                    str(applicant_dict.get('region', '')),
                    str(applicant_dict.get('city', '')),
                    str(applicant_dict.get('phone', '')),
                    str(applicant_dict.get('agitator_name', '')),
                ]
                if any(search_text in field.lower() for field in text_fields):
                    filtered_applicants.append(applicant)
            applicants = filtered_applicants

        self.table.blockSignals(True)
        self.table.setRowCount(len(applicants))

        id_to_row = {}

        # Категории для отображения
        category_map = {'м': 'м', 'ж': 'ж', 'всл': 'в/сл'}

        for row, applicant in enumerate(applicants):
            applicant_dict = dict(applicant)
            applicant_id = applicant_dict.get('id')
            id_to_row[applicant_id] = row

            phone = applicant_dict.get('phone', '')
            formatted_phone = self.format_phone_number(phone)

            # Отображение категории
            category = applicant_dict.get('category', '')
            category_display = category_map.get(category, category)

            # Статус
            status = applicant_dict.get('status', '')
            status_display = 'Поступает' if status == 'поступает' else 'Отказывается'

            # Дата добавления
            created_at = applicant_dict.get('created_at', '')
            if created_at:
                try:
                    # Парсим дату из SQLite
                    dt = datetime.fromisoformat(created_at.replace(' ', 'T'))
                    date_display = dt.strftime("%d.%m.%Y %H:%M")
                except:
                    date_display = str(created_at)[:16]
            else:
                date_display = ""

            items = [
                QTableWidgetItem(str(applicant_id)),
                QTableWidgetItem(applicant_dict.get('applicant_name', '')),
                QTableWidgetItem(applicant_dict.get('region', '')),
                QTableWidgetItem(applicant_dict.get('city', '')),
                QTableWidgetItem(category_display),
                QTableWidgetItem(formatted_phone),
                QTableWidgetItem(applicant_dict.get('education', '')),
                QTableWidgetItem(status_display),
                QTableWidgetItem(applicant_dict.get('document_status', '')),
                QTableWidgetItem(applicant_dict.get('agitator_department', '')),
                QTableWidgetItem(applicant_dict.get('agitator_name', '')),
                QTableWidgetItem(date_display),
            ]

            # Цветовая индикация
            if status == 'поступает':
                items[7].setBackground(QColor(230, 255, 230))
                items[7].setForeground(QColor(0, 100, 0))
            else:
                items[7].setBackground(QColor(255, 230, 230))
                items[7].setForeground(QColor(150, 0, 0))

            # Цвет для категории
            if category == 'м':
                items[4].setBackground(QColor(230, 240, 255))
            elif category == 'ж':
                items[4].setBackground(QColor(255, 230, 240))
            elif category == 'всл':
                items[4].setBackground(QColor(230, 255, 230))

            # Подсветка пустых полей
            for col, item in enumerate(items):
                if not item.text().strip() and col not in [0, 4]:  # ID и категория могут быть пустыми
                    item.setBackground(QColor(255, 255, 200))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, col, item)

        # Восстанавливаем выделение
        for applicant_id in selected_ids:
            if applicant_id in id_to_row:
                self.table.selectRow(id_to_row[applicant_id])

        self.table.resizeColumnsToContents()
        self.table.blockSignals(False)
        self.on_selection_changed()

    def init_settings_tab(self):
        """Инициализация вкладки настроек (только для админа)"""
        layout = QVBoxLayout()

        # Вкладки внутри настроек
        self.admin_tabs = QTabWidget()
        self.admin_tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #ddd;
                background-color: white;
                border-radius: 5px;
            }
            QTabBar::tab {
                background-color: #f8f9fa;
                padding: 10px 15px;
                margin-right: 2px;
                border: 1px solid #ddd;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        # Вкладка пользователей
        self.users_tab = QWidget()
        self.init_users_tab()
        self.admin_tabs.addTab(self.users_tab, QIcon("icons/users.png"), "Пользователи")

        # Вкладка подразделений (НОВАЯ)
        self.departments_tab = QWidget()
        self.init_departments_tab()
        self.admin_tabs.addTab(self.departments_tab, QIcon("icons/department.png"), "Подразделения")

        # Вкладка регионов (НОВАЯ)
        self.regions_tab = QWidget()
        self.init_regions_tab()
        self.admin_tabs.addTab(self.regions_tab, QIcon("icons/map.png"), "Регионы")

        # Вкладка образования
        self.education_tab = QWidget()
        self.init_education_tab()
        self.admin_tabs.addTab(self.education_tab, QIcon("icons/graduation.png"), "Образование")

        # Вкладка документов
        self.documents_tab = QWidget()
        self.init_documents_tab()
        self.admin_tabs.addTab(self.documents_tab, QIcon("icons/documents.png"), "Статусы документов")

        # Вкладка настройки дней когда можно редактировать
        self.schedule_tab = QWidget()
        self.init_schedule_tab()
        self.admin_tabs.addTab(self.schedule_tab, QIcon("icons/cloud.png"), "Расписание")
        # вкладка "Регионы"
        self.regions_tab = QWidget()
        self.init_regions_tab()
        self.admin_tabs.addTab(self.regions_tab, "Регионы")

        # Вкладка "Ответственные за регионы"
        self.department_regions_tab = QWidget()
        self.init_department_regions_tab()
        self.admin_tabs.addTab(self.department_regions_tab, "Ответственные за регионы")

        # Вкладка прав доступа
        self.permissions_tab = QWidget()
        self.init_permissions_tab()
        self.admin_tabs.addTab(self.permissions_tab, "Права доступа")

        layout.addWidget(self.admin_tabs)
        self.settings_tab.setLayout(layout)

    def init_department_regions_tab(self):
        """Инициализация вкладки управления ответственными за регионы"""
        layout = QVBoxLayout()

        # Панель управления
        controls_widget = QWidget()
        controls_layout = QHBoxLayout()

        self.department_combo_for_regions = QComboBox()
        self.department_combo_for_regions.addItems([d['name'] for d in self.db.get_departments()])
        self.department_combo_for_regions.currentIndexChanged.connect(self.refresh_department_regions)

        self.add_region_to_department_btn = QPushButton("Добавить регион")
        self.add_region_to_department_btn.clicked.connect(self.add_region_to_department)

        self.remove_region_from_department_btn = QPushButton("Удалить регион")
        self.remove_region_from_department_btn.clicked.connect(self.remove_region_from_department)

        controls_layout.addWidget(QLabel("Подразделение:"))
        controls_layout.addWidget(self.department_combo_for_regions)
        controls_layout.addWidget(self.add_region_to_department_btn)
        controls_layout.addWidget(self.remove_region_from_department_btn)
        controls_layout.addStretch()

        controls_widget.setLayout(controls_layout)
        layout.addWidget(controls_widget)

        # Список регионов подразделения
        self.department_regions_list = QListWidget()
        layout.addWidget(self.department_regions_list)

        self.department_regions_tab.setLayout(layout)
        self.refresh_department_regions()

    def refresh_department_regions(self):
        """Обновление списка регионов для выбранного подразделения"""
        department_name = self.department_combo_for_regions.currentText()
        if not department_name:
            return

        cursor = self.db.conn.cursor()
        cursor.execute('SELECT id FROM departments WHERE name = ?', (department_name,))
        department = cursor.fetchone()
        if not department:
            return

        department_id = department['id']
        regions = self.db.get_regions_for_department(department_id)
        self.department_regions_list.clear()
        for region in regions:
            self.department_regions_list.addItem(region['name'])

    def add_region_to_department(self):
        """Добавление региона к подразделению"""
        department_name = self.department_combo_for_regions.currentText()
        if not department_name:
            QMessageBox.warning(self, "Внимание", "Выберите подразделение!")
            return

        cursor = self.db.conn.cursor()
        cursor.execute('SELECT id FROM departments WHERE name = ?', (department_name,))
        department = cursor.fetchone()
        if not department:
            return

        department_id = department['id']
        regions = self.db.get_regions()
        region_name, ok = QInputDialog.getItem(self, "Добавить регион", "Выберите регион:", regions, 0, False)

        if ok and region_name:
            cursor.execute('SELECT id FROM regions WHERE name = ?', (region_name,))
            region = cursor.fetchone()
            if not region:
                return

            region_id = region['id']
            self.db.add_region_to_department(department_id, region_id)
            self.refresh_department_regions()

    def remove_region_from_department(self):
        """Удаление региона из подразделения"""
        current_item = self.department_regions_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Внимание", "Выберите регион для удаления!")
            return

        department_name = self.department_combo_for_regions.currentText()
        if not department_name:
            return

        cursor = self.db.conn.cursor()
        cursor.execute('SELECT id FROM departments WHERE name = ?', (department_name,))
        department = cursor.fetchone()
        if not department:
            return

        department_id = department['id']
        region_name = current_item.text()
        cursor.execute('SELECT id FROM regions WHERE name = ?', (region_name,))
        region = cursor.fetchone()
        if not region:
            return

        region_id = region['id']
        self.db.remove_region_from_department(department_id, region_id)
        self.refresh_department_regions()

    def init_regions_tab(self):
        """Инициализация вкладки управления регионами"""
        layout = QVBoxLayout()

        # Панель управления
        controls_widget = QWidget()
        controls_layout = QHBoxLayout()

        self.add_region_btn = QPushButton("Добавить регион")
        self.add_region_btn.clicked.connect(self.add_region)
        self.add_region_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
        """)

        self.delete_region_btn = QPushButton("Удалить регион")
        self.delete_region_btn.clicked.connect(self.delete_region)
        self.delete_region_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
        """)

        controls_layout.addWidget(self.add_region_btn)
        controls_layout.addWidget(self.delete_region_btn)
        controls_layout.addStretch()

        controls_widget.setLayout(controls_layout)
        layout.addWidget(controls_widget)

        # Список регионов
        self.regions_list = QListWidget()
        self.refresh_regions_list()
        layout.addWidget(self.regions_list)

        self.regions_tab.setLayout(layout)

    def refresh_regions_list(self):
        """Обновление списка регионов"""
        self.regions_list.clear()
        regions = self.db.get_regions()
        for region in regions:
            self.regions_list.addItem(region)

    def add_region(self):
        """Добавление региона"""
        text, ok = QInputDialog.getText(self, "Добавить регион", "Введите название региона:")
        if ok and text.strip():
            success = self.db.add_region(text.strip())
            if success:
                self.refresh_regions_list()
                QMessageBox.information(self, "Успех", "Регион добавлен!")
            else:
                QMessageBox.warning(self, "Ошибка", "Такой регион уже существует!")

    def delete_region(self):
        """Удаление региона"""
        current = self.regions_list.currentItem()
        if not current:
            QMessageBox.warning(self, "Внимание", "Выберите регион для удаления!")
            return

        reply = QMessageBox.question(
            self, "Подтверждение",
            f"Удалить '{current.text()}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            success = self.db.delete_region(current.text())
            if success:
                self.refresh_regions_list()
                QMessageBox.information(self, "Успех", "Регион удален!")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить регион!")

    def init_schedule_tab(self):
        """Инициализация вкладки расписания"""
        layout = QVBoxLayout()

        # Группа дней недели
        days_group = QGroupBox("Дни недели для добавления записей")
        days_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)

        days_layout = QVBoxLayout()

        # Чекбоксы для дней недели
        self.day_checkboxes = []
        days_names = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

        work_days = self.db.get_work_days()

        for i, day_name in enumerate(days_names, 1):
            checkbox = QCheckBox(day_name)
            checkbox.setChecked(i in work_days)
            self.day_checkboxes.append((i, checkbox))
            days_layout.addWidget(checkbox)

        days_group.setLayout(days_layout)
        layout.addWidget(days_group)

        # Кнопка сохранения
        save_btn = QPushButton("Сохранить настройки")
        save_btn.clicked.connect(self.save_work_days)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                padding: 10px;
                font-weight: bold;
            }
        """)
        layout.addWidget(save_btn)

        layout.addStretch()
        self.schedule_tab.setLayout(layout)

    def save_work_days(self):
        """Сохранение настроек дней недели"""
        selected_days = [day_num for day_num, cb in self.day_checkboxes if cb.isChecked()]
        if not selected_days:
            QMessageBox.warning(self, "Ошибка", "Выберите хотя бы один день!")
            return

        self.db.set_work_days(selected_days)
        QMessageBox.information(self, "Успех", "Настройки сохранены!")

    def init_education_tab(self):
        """Инициализация вкладки управления образованием"""
        layout = QVBoxLayout()

        # Панель управления
        controls_widget = QWidget()
        controls_layout = QHBoxLayout()

        self.add_edu_btn = QPushButton("Добавить")
        self.add_edu_btn.clicked.connect(self.add_education_type)
        self.add_edu_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
        """)

        self.delete_edu_btn = QPushButton("Удалить")
        self.delete_edu_btn.clicked.connect(self.delete_education_type)
        self.delete_edu_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
        """)

        controls_layout.addWidget(self.add_edu_btn)
        controls_layout.addWidget(self.delete_edu_btn)
        controls_layout.addStretch()

        controls_widget.setLayout(controls_layout)
        layout.addWidget(controls_widget)

        # Список образования
        self.education_list = QListWidget()
        self.refresh_education_list()
        layout.addWidget(self.education_list)

        self.education_tab.setLayout(layout)

    def init_documents_tab(self):
        """Инициализация вкладки управления статусами документов"""
        layout = QVBoxLayout()

        # Панель управления
        controls_widget = QWidget()
        controls_layout = QHBoxLayout()

        self.add_doc_btn = QPushButton("Добавить")
        self.add_doc_btn.clicked.connect(self.add_document_status)
        self.add_doc_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
        """)

        self.delete_doc_btn = QPushButton("Удалить")
        self.delete_doc_btn.clicked.connect(self.delete_document_status)
        self.delete_doc_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
        """)

        controls_layout.addWidget(self.add_doc_btn)
        controls_layout.addWidget(self.delete_doc_btn)
        controls_layout.addStretch()

        controls_widget.setLayout(controls_layout)
        layout.addWidget(controls_widget)

        # Список статусов документов
        self.documents_list = QListWidget()
        self.refresh_documents_list()
        layout.addWidget(self.documents_list)

        self.documents_tab.setLayout(layout)

    def refresh_education_list(self):
        """Обновление списка образования"""
        self.education_list.clear()
        education_types = self.db.get_education_types()
        for edu in education_types:
            self.education_list.addItem(edu)

    def refresh_documents_list(self):
        """Обновление списка статусов документов"""
        self.documents_list.clear()
        doc_statuses = self.db.get_document_statuses()
        for doc in doc_statuses:
            self.documents_list.addItem(doc)

    def add_education_type(self):
        """Добавление типа образования"""
        text, ok = QInputDialog.getText(self, "Добавить тип образования", "Введите название:")
        if ok and text.strip():
            success = self.db.add_education_type(text.strip())
            if success:
                self.refresh_education_list()
                QMessageBox.information(self, "Успех", "Тип образования добавлен!")
            else:
                QMessageBox.warning(self, "Ошибка", "Такой тип уже существует!")

    def delete_education_type(self):
        """Удаление типа образования"""
        current = self.education_list.currentItem()
        if not current:
            QMessageBox.warning(self, "Внимание", "Выберите тип для удаления!")
            return

        reply = QMessageBox.question(
            self, "Подтверждение",
            f"Удалить '{current.text()}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            success = self.db.delete_education_type(current.text())
            if success:
                self.refresh_education_list()
                QMessageBox.information(self, "Успех", "Тип образования удален!")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить!")

    def add_document_status(self):
        """Добавление статуса документов"""
        text, ok = QInputDialog.getText(self, "Добавить статус документов", "Введите название:")
        if ok and text.strip():
            success = self.db.add_document_status(text.strip())
            if success:
                self.refresh_documents_list()
                QMessageBox.information(self, "Успех", "Статус документов добавлен!")
            else:
                QMessageBox.warning(self, "Ошибка", "Такой статус уже существует!")

    def delete_document_status(self):
        """Удаление статуса документов"""
        current = self.documents_list.currentItem()
        if not current:
            QMessageBox.warning(self, "Внимание", "Выберите статус для удаления!")
            return

        reply = QMessageBox.question(
            self, "Подтверждение",
            f"Удалить '{current.text()}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            success = self.db.delete_document_status(current.text())
            if success:
                self.refresh_documents_list()
                QMessageBox.information(self, "Успех", "Статус документов удален!")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить!")

    def refresh_departments(self):
        """Обновление списка подразделений"""
        departments = self.db.get_all_departments_with_heads()

        self.departments_table.setRowCount(len(departments))

        for row, dept in enumerate(departments):
            dept_dict = dict(dept)

            items = [
                QTableWidgetItem(str(dept_dict.get('id', ''))),
                QTableWidgetItem(dept_dict.get('name', '')),
                QTableWidgetItem(self.get_department_type_text(dept_dict.get('type', 'department'))),
                QTableWidgetItem(dept_dict.get('head_name', 'Не назначен')),
            ]

            for col, item in enumerate(items):
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.departments_table.setItem(row, col, item)

        self.departments_table.resizeColumnsToContents()

    def get_department_type_text(self, dept_type):
        """Преобразование типа подразделения в читаемый текст"""
        types = {
            'faculty': 'Факультет',
            'department': 'Кафедра',
            'group': 'Группа',
            'root': 'Корневое'
        }
        return types.get(dept_type, dept_type)

    def init_departments_tab(self):
        """Инициализация вкладки управления подразделениями"""
        layout = QVBoxLayout()

        # Панель управления
        controls_widget = QWidget()
        controls_layout = QHBoxLayout()

        self.add_dept_btn = QPushButton("Добавить подразделение")
        self.add_dept_btn.clicked.connect(self.add_department)
        self.add_dept_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)

        self.edit_dept_btn = QPushButton("Редактировать")
        self.edit_dept_btn.clicked.connect(self.edit_department)
        self.edit_dept_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)

        self.delete_dept_btn = QPushButton("Удалить")
        self.delete_dept_btn.clicked.connect(self.delete_department)
        self.delete_dept_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)

        controls_layout.addWidget(self.add_dept_btn)
        controls_layout.addWidget(self.edit_dept_btn)
        controls_layout.addWidget(self.delete_dept_btn)
        controls_layout.addStretch()

        controls_widget.setLayout(controls_layout)
        layout.addWidget(controls_widget)

        # Таблица подразделений
        self.departments_table = QTableWidget()
        self.departments_table.setColumnCount(4)
        self.departments_table.setHorizontalHeaderLabels([
            "ID", "Название", "Тип", "Начальник"
        ])
        self.departments_table.setColumnHidden(0, True)  # Скрываем ID
        self.departments_table.horizontalHeader().setStretchLastSection(True)
        self.departments_table.setAlternatingRowColors(True)
        self.departments_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.departments_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        layout.addWidget(self.departments_table)

        self.departments_tab.setLayout(layout)
        self.refresh_departments()

    def refresh_departments(self):
        """Обновление списка подразделений"""
        departments = self.db.get_all_departments_with_heads()

        self.departments_table.setRowCount(len(departments))

        for row, dept in enumerate(departments):
            dept_dict = dict(dept)

            items = [
                QTableWidgetItem(str(dept_dict.get('id', ''))),
                QTableWidgetItem(dept_dict.get('name', '')),
                QTableWidgetItem(self.get_department_type_text(dept_dict.get('type', 'department'))),
                QTableWidgetItem(dept_dict.get('head_name', '')),
            ]

            for col, item in enumerate(items):
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.departments_table.setItem(row, col, item)

        self.departments_table.resizeColumnsToContents()

    def get_department_type_text(self, dept_type):
        """Преобразование типа подразделения в читаемый текст"""
        types = {
            'faculty': 'Факультет',
            'department': 'Кафедра',
            'group': 'Группа',
            'root': 'Корневое'
        }
        return types.get(dept_type, dept_type)

    def add_department(self):
        """Добавление подразделения"""
        dialog = DepartmentDialog(db=self.db, parent=self)
        if dialog.exec():
            data = dialog.get_data()

            # Добавляем подразделение
            dept_id = self.db.add_department(
                data['name'],
                data['type'],
                data.get('parent_id')
            )

            if dept_id:
                # Если выбран начальник, назначаем его
                if data.get('head_user_id'):
                    self.db.set_department_head(dept_id, data['head_user_id'])

                QMessageBox.information(self, "Успех", "Подразделение добавлено!")
                self.refresh_departments()
                self.load_departments_for_combo()  # Обновляем комбобоксы в других местах
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось добавить подразделение!")

    def edit_department(self):
        """Редактирование подразделения"""
        selected_rows = self.departments_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Внимание", "Выберите подразделение!")
            return

        row = selected_rows[0].row()
        dept_id = int(self.departments_table.item(row, 0).text())

        # Получаем данные подразделения
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT * FROM departments WHERE id = ?', (dept_id,))
        dept_data = dict(cursor.fetchone())

        dialog = DepartmentDialog(dept_data, self.db, self)
        if dialog.exec():
            data = dialog.get_data()

            # Обновляем подразделение
            success = self.db.update_department(dept_id, data)

            if success:
                # Обновляем начальника
                self.db.set_department_head(dept_id, data.get('head_user_id'))

                QMessageBox.information(self, "Успех", "Подразделение обновлено!")
                self.refresh_departments()
                self.load_departments_for_combo()
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось обновить подразделение!")

    def delete_department(self):
        """Удаление подразделения"""
        selected_rows = self.departments_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Внимание", "Выберите подразделение!")
            return

        row = selected_rows[0].row()
        dept_name = self.departments_table.item(row, 1).text()
        dept_id = int(self.departments_table.item(row, 0).text())

        # Проверяем, есть ли связанные данные
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT COUNT(*) as count FROM applicants WHERE agitator_department = ?', (dept_name,))
        applicants_count = cursor.fetchone()['count']

        message = f"Вы уверены, что хотите удалить подразделение '{dept_name}'?"
        if applicants_count > 0:
            message += f"\n\nВНИМАНИЕ: К этому подразделению привязано {applicants_count} абитуриентов!\nИх данные останутся, но связь с подразделением будет потеряна."

        reply = QMessageBox.question(
            self, "Подтверждение",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Сначала удаляем права доступа
            cursor.execute('DELETE FROM user_department_permissions WHERE department_id = ?', (dept_id,))
            # Затем удаляем подразделение
            success = self.db.delete_department(dept_id)

            if success:
                QMessageBox.information(self, "Успех", "Подразделение удалено!")
                self.refresh_departments()
                self.load_departments_for_combo()
            else:
                QMessageBox.critical(self, "Ошибка", "Не удалось удалить подразделение!")

    def load_departments_for_combo(self):
        """Обновление списка подразделений во всех комбобоксах"""
        # Обновляем комбобокс в UserDialog при следующем открытии
        # Здесь можно обновить глобальные ссылки, если нужно
        pass

    def init_users_tab(self):
        """Инициализация вкладки пользователей"""
        layout = QVBoxLayout()

        # Панель управления
        controls_widget = QWidget()
        controls_layout = QHBoxLayout()

        self.add_user_btn = QPushButton("Добавить пользователя")
        self.add_user_btn.clicked.connect(self.add_user)

        self.edit_user_btn = QPushButton("Редактировать")
        self.edit_user_btn.clicked.connect(self.edit_user)

        self.delete_user_btn = QPushButton("Удалить")
        self.delete_user_btn.clicked.connect(self.delete_user)

        self.search_user_input = QLineEdit()
        self.search_user_input.setPlaceholderText("Поиск пользователей...")
        self.search_user_input.textChanged.connect(self.refresh_users)
        self.search_user_input.setMinimumWidth(250)

        controls_layout.addWidget(self.add_user_btn)
        controls_layout.addWidget(self.edit_user_btn)
        controls_layout.addWidget(self.delete_user_btn)
        controls_layout.addStretch()
        controls_layout.addWidget(self.search_user_input)

        controls_widget.setLayout(controls_layout)
        layout.addWidget(controls_widget)

        # Таблица пользователей (добавляем колонку "Начальник")
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(8)
        self.users_table.setHorizontalHeaderLabels([
            "ID", "Логин", "ФИО", "Роль", "Подразделение", "Должность", "Звание", "Начальник"
        ])
        self.users_table.setColumnHidden(0, True)  # Скрываем ID

        # Настройка таблицы
        self.users_table.horizontalHeader().setStretchLastSection(True)
        self.users_table.setAlternatingRowColors(True)
        self.users_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.users_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        layout.addWidget(self.users_table)

        self.users_tab.setLayout(layout)
        self.refresh_users()

    @staticmethod
    def format_phone_number(phone):
        """Форматирование номера телефона в стандартный формат"""
        if not phone:
            return ""

        # Удаляем все нецифровые символы
        digits = ''.join(filter(str.isdigit, str(phone)))

        # Проверяем длину
        if len(digits) < 10:
            return phone  # Возвращаем как есть, если номер слишком короткий

        # Определяем код страны
        if digits.startswith('8') and len(digits) == 11:
            # Российский номер в формате 8XXXXXXXXXX
            digits = '7' + digits[1:]
        elif digits.startswith('7') and len(digits) == 11:
            pass  # Уже в правильном формате
        elif len(digits) == 10:
            # Номер без кода страны
            digits = '7' + digits

        # Форматируем в стандартный вид
        if len(digits) >= 11:
            return f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:]}"

        return phone

    def init_permissions_tab(self):
        """Инициализация вкладки прав доступа"""
        layout = QVBoxLayout()

        # Панель управления
        controls_widget = QWidget()
        controls_layout = QHBoxLayout()

        # Выбор пользователя
        user_label = QLabel('Пользователь:')
        self.user_combo = QComboBox()
        self.user_combo.currentIndexChanged.connect(self.refresh_permissions)

        self.add_permission_btn = QPushButton('Добавить права')
        self.add_permission_btn.clicked.connect(self.add_permission)

        # Добавление виджетов
        controls_layout.addWidget(user_label)
        controls_layout.addWidget(self.user_combo)
        controls_layout.addWidget(self.add_permission_btn)
        controls_layout.addStretch()

        controls_widget.setLayout(controls_layout)
        layout.addWidget(controls_widget)

        # Таблица прав доступа
        self.permissions_table = QTableWidget()
        self.permissions_table.setColumnCount(4)
        self.permissions_table.setHorizontalHeaderLabels([
            'Пользователь', 'Тип права', 'Курс'
        ])

        # Настройка таблицы
        self.permissions_table.horizontalHeader().setStretchLastSection(True)
        self.permissions_table.setAlternatingRowColors(True)
        self.permissions_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.permissions_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        layout.addWidget(self.permissions_table)

        self.permissions_tab.setLayout(layout)

        # Загрузка пользователей для комбобокса
        self.load_users_for_combo()

    def load_users_for_combo(self):
        """Загрузка пользователей в комбобокс"""
        self.user_combo.clear()
        self.user_combo.addItem('-- Выберите пользователя --', None)

        users = self.db.get_all_users()
        for user in users:
            user_dict = dict(user)
            if user_dict['role'] != 'admin':  # Не показываем администраторов
                self.user_combo.addItem(f"{user_dict['username']} ({user_dict['full_name']})", user_dict['id'])

    def refresh_users(self):
        """Обновление списка пользователей"""
        users = self.db.get_all_users()
        search_text = self.search_user_input.text().lower().strip()

        if search_text:
            filtered_users = []
            for user in users:
                user_dict = dict(user)
                if (search_text in user_dict['username'].lower() or
                        search_text in user_dict['full_name'].lower()):
                    filtered_users.append(user)
            users = filtered_users

        self.users_table.setRowCount(len(users))

        role_map = {'admin': 'Администратор', 'user': 'Пользователь'}

        for row, user in enumerate(users):
            user_dict = dict(user)

            items = [
                QTableWidgetItem(str(user_dict.get('id', ''))),
                QTableWidgetItem(user_dict.get('username', '')),
                QTableWidgetItem(user_dict.get('full_name', '')),
                QTableWidgetItem(role_map.get(user_dict.get('role', ''), user_dict.get('role', ''))),
                QTableWidgetItem(user_dict.get('department_name', '')),
                QTableWidgetItem(user_dict.get('position', '')),
                QTableWidgetItem(user_dict.get('rank', '')),
                QTableWidgetItem("Да" if user_dict.get('is_head', False) else "Нет"),
            ]

            for col, item in enumerate(items):
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.users_table.setItem(row, col, item)

        self.users_table.resizeColumnsToContents()

    def refresh_permissions(self):
        """Обновление списка прав доступа"""
        user_id = self.user_combo.currentData()

        if not user_id:
            self.permissions_table.setRowCount(0)
            return

        permissions = self.db.get_user_permissions(user_id)

        # Заполнение таблицы
        self.permissions_table.setRowCount(len(permissions))

        for row, permission in enumerate(permissions):
            perm_dict = dict(permission)

            items = [
                QTableWidgetItem(f"{perm_dict.get('full_name', '')}"),
                QTableWidgetItem(perm_dict.get('permission_type', '')),
                QTableWidgetItem(perm_dict.get('course', '')),
                # QTableWidgetItem(perm_dict.get('faculty', ''))
            ]

            for col, item in enumerate(items):
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.permissions_table.setItem(row, col, item)

        self.permissions_table.resizeColumnsToContents()

    # В main.py, исправьте метод add_user:

    def add_user(self):
        """Добавление нового пользователя"""
        dialog = UserDialog(db=self.db, parent=self)
        if dialog.exec():
            data = dialog.get_data()
            if data is None:
                return

            # Добавляем пользователя с обработкой ошибок
            try:
                user_id = self.db.add_user(
                    data['username'],
                    data['password'],
                    data['full_name'],
                    data['role'],
                    data.get('department_id'),
                    data.get('position', ''),
                    data.get('rank', ''),
                    data.get('is_head', False)
                )

                if user_id:
                    # Добавляем права доступа (только для обычных пользователей, не начальников)
                    if data['role'] != 'admin' and not data.get('is_head', False):
                        for dept_name in data.get('permissions', []):
                            cursor = self.db.conn.cursor()
                            cursor.execute('SELECT id FROM departments WHERE name = ?', (dept_name,))
                            result = cursor.fetchone()
                            if result:
                                self.db.add_user_department_permission(user_id, result['id'], True, False)

                    QMessageBox.information(self, 'Успех', 'Пользователь успешно добавлен!')
                    self.refresh_users()
                    self.load_users_for_combo()
                else:
                    QMessageBox.critical(self, 'Ошибка',
                                         'Не удалось добавить пользователя. Возможно, логин уже существует.')
            except Exception as e:
                QMessageBox.critical(self, 'Ошибка', f'Произошла ошибка: {str(e)}')

    def edit_user(self):
        """Редактирование выбранного пользователя"""
        selected_rows = self.users_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Внимание', 'Выберите пользователя для редактирования!')
            return

        row = selected_rows[0].row()
        user_id_item = self.users_table.item(row, 0)
        if not user_id_item:
            return
        user_id = int(user_id_item.text())

        # Получение данных пользователя из БД
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        user_data = dict(cursor.fetchone())

        # Добавляем название подразделения
        if user_data.get('department_id'):
            cursor.execute('SELECT name FROM departments WHERE id = ?', (user_data['department_id'],))
            dept = cursor.fetchone()
            if dept:
                user_data['department_name'] = dept['name']

        dialog = UserDialog(user_data=user_data, db=self.db, parent=self)
        if dialog.exec():
            new_data = dialog.get_data()
            if new_data is None:
                return

            # Обновляем пользователя
            success = self.db.update_user(user_data['id'], new_data)

            if success:
                # Обновляем права доступа
                # Сначала удаляем старые
                cursor.execute('DELETE FROM user_department_permissions WHERE user_id = ?', (user_data['id'],))

                # Добавляем новые
                for dept_name in new_data['permissions']:
                    cursor.execute('SELECT id FROM departments WHERE name = ?', (dept_name,))
                    result = cursor.fetchone()
                    if result:
                        self.db.add_user_department_permission(user_data['id'], result['id'], True, False)

                QMessageBox.information(self, 'Успех', 'Данные пользователя обновлены!')
                self.refresh_users()
                self.load_users_for_combo()
            else:
                QMessageBox.critical(self, 'Ошибка', 'Не удалось обновить данные пользователя.')

    def delete_user(self):
        """Удаление выбранного пользователя"""
        selected_rows = self.users_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Внимание', 'Выберите пользователя для удаления!')
            return

        row = selected_rows[0].row()
        user_id_item = self.users_table.item(row, 0)
        if not user_id_item:
            return
        user_id = int(user_id_item.text())
        username = self.users_table.item(row, 1).text()

        # Нельзя удалить самого себя
        if user_id == self.user_data['id']:
            QMessageBox.warning(self, 'Ошибка', 'Вы не можете удалить самого себя!')
            return

        reply = QMessageBox.question(
            self, 'Подтверждение',
            f'Вы уверены, что хотите удалить пользователя "{username}"?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            success = self.db.delete_user(user_id)

            if success:
                QMessageBox.information(self, 'Успех', 'Пользователь успешно удален!')
                self.refresh_users()
                self.load_users_for_combo()
            else:
                QMessageBox.critical(self, 'Ошибка', 'Не удалось удалить пользователя.')

    def add_permission(self):
        """Добавление нового права доступа"""
        user_id = self.user_combo.currentData()

        if not user_id:
            QMessageBox.warning(self, 'Внимание', 'Выберите пользователя!')
            return

        user_name = self.user_combo.currentText()
        dialog = PermissionDialog(user_id, user_name, self)

        if dialog.exec():
            data = dialog.get_data()
            permission_id = self.db.add_permission(
                data['user_id'], data['permission_type'], None,
                data['course']
            )

            if permission_id:
                QMessageBox.information(self, 'Успех', 'Права доступа успешно добавлено!')
                self.refresh_permissions()
            else:
                QMessageBox.critical(self, 'Ошибка', 'Не удалось добавить права доступа.')

    def check_can_add(self):
        """Проверка, можно ли добавлять сегодня"""
        from datetime import datetime
        current_day = datetime.now().weekday() + 1  # пн=1, вс=7
        work_days = self.db.get_work_days()
        return current_day in work_days

    def add_applicant(self):
        """Добавление нового абитуриента"""
        if not self.check_can_add():
            QMessageBox.warning(self, "Внимание",
                                "В выходные дни добавление записей запрещено!\n"
                                "Пожалуйста, обратитесь к администратору.")
            return

        dialog = ApplicantDialog(
            applicant_data=None,
            user_role=self.user_data['role'],
            db=self.db,
            parent=self
        )
        if dialog.exec():
            data = dialog.get_data()

            # Проверка обязательных полей
            if not data['applicant_name']:
                QMessageBox.warning(self, 'Ошибка', 'ФИО абитуриента обязательно!')
                return

            if not data['agitator_name']:
                QMessageBox.warning(self, 'Ошибка', 'ФИО агитатора обязательно!')
                return

            # Проверка на дубликат
            if self.db.check_duplicate_applicant(data['applicant_name'], data['phone']):
                reply = QMessageBox.question(
                    self, 'Дубликат',
                    'Абитуриент с таким ФИО и телефоном уже существует!\n'
                    'Вы всё равно хотите добавить?',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return

            # Добавляем в БД
            self.db.add_applicant(self.user_data['id'], data)

            # Обновляем таблицу и статистику
            self.refresh_data()
            self.stats_tab.update_statistics()

            QMessageBox.information(self, 'Успех', 'Абитуриент успешно добавлен!')

    def edit_applicant(self):
        """Редактирование выбранного абитуриента"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Внимание', 'Выберите запись для редактирования!')
            return

        row = selected_rows[0].row()
        applicant_id = int(self.table.item(row, 0).text())

        # Получение данных абитуриента из БД
        cursor = self.db.conn.cursor()
        cursor.execute('SELECT * FROM applicants WHERE id = ?', (applicant_id,))
        applicant_data = dict(cursor.fetchone())

        dialog = ApplicantDialog(
            applicant_data=applicant_data,
            user_role=self.user_data['role'],
            db=self.db,
            parent=self
        )
        if dialog.exec():
            data = dialog.get_data()
            self.db.update_applicant(applicant_id, data)
            self.refresh_data()
            self.stats_tab.update_statistics()
            QMessageBox.information(self, 'Успех', 'Данные абитуриента обновлены!')

    def delete_applicant(self):
        """Удаление выбранного абитуриента"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Внимание', 'Выберите запись для удаления!')
            return

        reply = QMessageBox.question(
            self, 'Подтверждение',
            'Вы уверены, что хотите удалить выбранную запись?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            for index in selected_rows:
                row = index.row()
                applicant_id = int(self.table.item(row, 0).text())
                self.db.delete_applicant(applicant_id)

            self.refresh_data()
            self.stats_tab.update_statistics()

    def import_from_excel(self):
        """Импорт данных из Excel файла"""
        dialog = ImportDialog(self.db, self.user_data['id'], self)
        if dialog.exec():
            # Обновляем данные после импорта
            self.refresh_data()
            self.stats_tab.update_statistics()
            QMessageBox.information(self, "Успех", "Данные успешно импортированы и обновлены!")

    @staticmethod
    def check_file_access(file_path):
        """Проверка доступности файла"""
        try:
            import time
            # Пробуем открыть файл несколько раз
            for attempt in range(5):
                try:
                    with open(file_path, 'rb') as f:
                        f.read(1)
                    return True
                except PermissionError:
                    time.sleep(0.1)
                    continue
            return False
        except Exception:
            return False

    def import_excel_data(self, file_path, excel_password, selected_sheets):
        """Импорт данных из Excel файла"""
        temp_file_path = None
        try:
            # Если указан пароль Excel, пробуем расшифровать
            if excel_password:
                try:
                    import msoffcrypto
                    import tempfile

                    with open(file_path, "rb") as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password=excel_password)

                        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                        temp_file_path = temp_file.name
                        office_file.decrypt(temp_file)
                        temp_file.close()

                        file_path = temp_file_path

                except Exception as e:
                    return False, f"Неверный пароль Excel файла или ошибка расшифровки: {str(e)}"

            try:
                # Определяем движок для чтения Excel
                engine = None
                if file_path.endswith('.xlsx'):
                    engine = 'openpyxl'
                elif file_path.endswith('.xls'):
                    engine = 'xlrd'  # Для старых .xls файлов
                else:
                    engine = 'openpyxl'  # по умолчанию

                # Пробуем прочитать файл
                try:
                    xls = pd.ExcelFile(file_path, engine=engine)
                except Exception as e:
                    # Если не получилось с выбранным движком, пробуем другой
                    if engine == 'openpyxl':
                        engine = 'xlrd'
                    else:
                        engine = 'openpyxl'

                    try:
                        xls = pd.ExcelFile(file_path, engine=engine)
                    except Exception as e2:
                        return False, f"Не удалось прочитать файл. Возможно, файл поврежден или имеет неподдерживаемый формат. Ошибки: {str(e)}, {str(e2)}"

                imported_count = 0
                duplicate_count = 0
                skipped_count = 0

                # Проверяем доступные листы
                available_sheets = xls.sheet_names

                for sheet in selected_sheets:
                    if sheet not in available_sheets:
                        return False, f"Лист '{sheet}' не найден в файле"

                for sheet_name in selected_sheets:
                    try:
                        # Читаем лист
                        df = pd.read_excel(xls, sheet_name=sheet_name, header=0)

                        # Определяем курс из названия листа
                        course_from_sheet = '1 курс'
                        for course_num in ['1', '2', '3', '4', '5']:
                            if f'{course_num} курс' in sheet_name.lower():
                                course_from_sheet = f'{course_num} курс'
                                break
                            elif f'курс {course_num}' in sheet_name.lower():
                                course_from_sheet = f'{course_num} курс'
                                break

                        # Импортируем данные
                        for index, row in df.iterrows():
                            # Пропускаем пустые строки
                            if row.isnull().all():
                                continue

                            # Получаем данные
                            applicant_data = self.extract_applicant_data(row, df, course_from_sheet)

                            if applicant_data:
                                # Проверяем дубликат
                                if self.check_duplicate(applicant_data):
                                    duplicate_count += 1
                                    continue

                                # Добавляем в базу
                                try:
                                    self.db.add_applicant(self.user_data['id'], applicant_data)
                                    imported_count += 1
                                except Exception as e:
                                    skipped_count += 1
                                    print(f"Ошибка добавления абитуриента: {e}")
                            else:
                                skipped_count += 1

                    except Exception as e:
                        error_msg = f"Ошибка импорта листа {sheet_name}: {e}"
                        print(error_msg)
                        import traceback
                        traceback.print_exc()
                        # Продолжаем с другими листами
                        continue

                result_message = f"""
                Импорт завершен!
                Всего импортировано: {imported_count}
                Пропущено дубликатов: {duplicate_count}
                Пропущено других ошибок: {skipped_count}
                Листы: {', '.join(selected_sheets)}
                """

                return True, result_message

            except Exception as e:
                error_msg = f"Ошибка чтения Excel файла: {str(e)}"
                print(error_msg)
                import traceback
                traceback.print_exc()
                return False, error_msg

            finally:
                # ЗАКРЫВАЕМ ExcelFile если он был открыт
                if 'xls' in locals():
                    xls.close()  # Явно закрываем файл

                # Удаляем временный файл если он был создан
                if temp_file_path and os.path.exists(temp_file_path):
                    try:
                        # Даем системе время освободить файл
                        import time
                        time.sleep(0.1)

                        # Пробуем удалить файл несколько раз
                        for attempt in range(3):
                            try:
                                os.unlink(temp_file_path)
                                break
                            except PermissionError:
                                time.sleep(0.1)
                                continue
                            except Exception as e:
                                print(f"Ошибка удаления временного файла: {e}")
                                break
                    except Exception as e:
                        print(f"Ошибка при удалении временного файла: {e}")

        except Exception as e:
            error_msg = f"Общая ошибка импорта: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg

    def extract_applicant_data(self, row, df, course):
        """Извлечение данных абитуриента из строки Excel"""
        try:
            # Ищем ФИО абитуриента
            applicant_name = ''
            for col in df.columns:
                col_str = str(col)
                if 'абитуриент' in col_str.lower():
                    value = row[col]
                    if pd.notna(value):
                        applicant_name = str(value).strip()
                        break

            if not applicant_name:
                return None

            # Собираем данные
            applicant_data = {
                'study_group': self._get_cell_value(row, ['уч.гр.', 'Уч. группа', 'учебная группа', 'Учебная группа',
                                                          'Группа']),
                'rank': self._get_cell_value(row, ['В.зв', 'Звание', 'воинское звание', 'Воинское звание', 'звание'],
                                             'ряд.'),
                'student_name': self._get_cell_value(row, ['ФИО', 'ФИО студента', 'Студент', 'фио', 'Студент ФИО']),
                'region': self._get_cell_value(row,
                                               ['Субъект РФ', 'Регион', 'область', 'Субъект_РФ', 'Регион (область)']),
                'city': self._get_cell_value(row, ['Населённый пункт', 'Город', 'город', 'Населенный пункт',
                                                   'Город/населенный пункт']),
                'category': self._get_cell_value(row, ['категория', 'Категория', 'пол', 'Пол', 'Кат.'], 'муж'),
                'applicant_name': applicant_name,
                'phone': self.format_imported_phone(
                self._get_cell_value(row, ['Телефон', 'телефон', 'контактный телефон', 'Телефон абитуриента', 'Тел.'])),
                'status': self._get_cell_value(row, ['Статус', 'статус', 'Статус поступления', 'Статус абитуриента'],
                                               'поступает'),
                'document_status': self._get_cell_value(row, ['Состояние личного дела на поступление', 'Документы',
                                                              'документы', 'Статус документов']),
                'notes': self._get_cell_value(row, ['Примечание', 'примечание', 'комментарий', 'Комментарий', 'Прим.']),
                'course': course,
                # 'faculty': self._get_cell_value(row, ['Факультет', 'факультет', 'Факультет/отделение', 'Отделение'])
            }

            # Нормализация статуса
            status = applicant_data['status'].lower()
            if 'поступает' in status or status == '1' or status == '1)поступает' or '1)' in status:
                applicant_data['status'] = '1)поступает'
            elif 'не поступает' in status or status == '0' or status == '2)не поступает' or '2)' in status or 'не' in status:
                applicant_data['status'] = '2)не поступает'
            else:
                applicant_data['status'] = '1)поступает'

            # Нормализация статуса документов
            doc_status = applicant_data['document_status'].lower()
            if 'формируется' in doc_status or doc_status == '1' or doc_status == '1)формируется' or '1)' in doc_status:
                applicant_data['document_status'] = '1)Формируется в военкомате'
            elif 'отправлено' in doc_status or doc_status == '2' or doc_status == '2)отправлено' or '2)' in doc_status:
                applicant_data['document_status'] = '2)Отправлено в ВА ВКО'
            elif 'ва вко' in doc_status or doc_status == '3' or doc_status == '3)в ва вко' or '3)' in doc_status:
                applicant_data['document_status'] = '3)В ВА ВКО'
            else:
                applicant_data['document_status'] = ''

            return applicant_data

        except Exception as e:
            print(f"Ошибка извлечения данных: {e}")
            import traceback
            traceback.print_exc()
            return None

    @staticmethod
    def _get_cell_value(row, possible_columns, default=''):
        """Получение значения ячейки из возможных колонок"""
        for col in possible_columns:
            if col in row and pd.notna(row[col]):
                return str(row[col]).strip()
        return default

    @staticmethod
    def format_imported_phone(phone):
        """Форматирование номера телефона при импорте"""
        if not phone:
            return ""

        # Удаляем все нецифровые символы
        digits = ''.join(filter(str.isdigit, str(phone)))

        if not digits:
            return phone

        # Приводим к стандартному формату
        if digits.startswith('8') and len(digits) == 11:
            # 8XXXXXXXXXX -> 7XXXXXXXXXX
            return '7' + digits[1:]
        elif len(digits) == 10:
            # XXXXXXXXXX -> 7XXXXXXXXXX
            return '7' + digits
        elif digits.startswith('7') and len(digits) == 11:
            return digits

        return phone

    def check_duplicate(self, applicant_data):
        """Проверка на дубликат (все пользователи)"""
        cursor = self.db.conn.cursor()
        cursor.execute('''
            SELECT COUNT(*) FROM applicants 
            WHERE applicant_name = ?
            AND phone = ?
            AND course = ?
        ''', (applicant_data['applicant_name'],
              applicant_data['phone'],
              applicant_data['course']))
        count = cursor.fetchone()[0]
        return count > 0

    def export_data(self):
        """Экспорт данных в Excel"""
        # Выбор файла для сохранения
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Сохранить данные',
            f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            'Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)'
        )

        if not file_path:
            return

        # Показываем прогресс
        progress = QProgressDialog("Экспорт данных...", "Отмена", 0, 100, self)
        progress.setWindowTitle("Экспорт")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()
        QApplication.processEvents()

        try:
            # Получаем данные для экспорта
            if file_path.endswith('.xlsx'):
                success = self.export_to_excel(file_path, progress)
            elif file_path.endswith('.csv'):
                success = self.export_to_csv(file_path, progress)
            else:
                # По умолчанию Excel
                if not file_path.endswith('.xlsx'):
                    file_path += '.xlsx'
                success = self.export_to_excel(file_path, progress)

            progress.setValue(100)

            if success:
                QMessageBox.information(self, 'Успех', f'Данные успешно экспортированы в:\n{file_path}')
            else:
                QMessageBox.critical(self, 'Ошибка', 'Не удалось экспортировать данные!')

        except Exception as e:
            progress.close()
            QMessageBox.critical(self, 'Ошибка', f'Ошибка при экспорте: {str(e)}')
            import traceback
            traceback.print_exc()

    def export_to_excel(self, file_path, progress):
        """Экспорт в Excel с форматированием"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            progress.setValue(10)
            QApplication.processEvents()

            # Получаем данные из таблицы
            data = []
            headers = []

            # Получаем заголовки (видимые колонки)
            for col in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(col)
                if header and not self.table.isColumnHidden(col):
                    headers.append(header.text())

            progress.setValue(30)
            QApplication.processEvents()

            # Получаем данные из строк
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    if not self.table.isColumnHidden(col):
                        item = self.table.item(row, col)
                        row_data.append(item.text() if item else '')
                data.append(row_data)

            progress.setValue(50)
            QApplication.processEvents()

            # Создаем DataFrame
            df = pd.DataFrame(data, columns=headers)

            # Добавляем строку с итогами
            total_row = ['ИТОГО:']
            for col in range(1, len(headers)):
                if headers[col] in ['Категория', 'Статус', 'Документы']:
                    # Для текстовых полей считаем количество непустых
                    col_data = [row[col] for row in data if row[col] and row[col].strip()]
                    total_row.append(str(len(col_data)))
                else:
                    # Для числовых полей (если есть)
                    total_row.append('')
            df.loc[len(df)] = total_row

            progress.setValue(70)
            QApplication.processEvents()

            # Сохраняем в Excel
            df.to_excel(file_path, index=False, sheet_name='Абитуриенты')

            # Форматируем Excel файл
            wb = load_workbook(file_path)
            ws = wb.active

            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            total_font = Font(bold=True, color="FFFFFF")
            total_fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Применяем стили к заголовкам
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

                # Автоширина колонок
                max_length = len(str(cell.value))
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                ws.column_dimensions[chr(64 + col)].width = min(max_length + 2, 50)

            # Стиль для строки итогов
            total_row_num = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=total_row_num, column=col)
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Сохраняем
            wb.save(file_path)

            progress.setValue(100)
            return True

        except Exception as e:
            print(f"Ошибка экспорта в Excel: {e}")
            import traceback
            traceback.print_exc()
            return False

    def export_to_csv(self, file_path, progress):
        """Экспорт в CSV"""
        try:
            import pandas as pd

            progress.setValue(20)
            QApplication.processEvents()

            # Получаем данные из таблицы
            data = []
            headers = []

            # Получаем заголовки (видимые колонки)
            for col in range(self.table.columnCount()):
                header = self.table.horizontalHeaderItem(col)
                if header and not self.table.isColumnHidden(col):
                    headers.append(header.text())

            progress.setValue(40)
            QApplication.processEvents()

            # Получаем данные из строк
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    if not self.table.isColumnHidden(col):
                        item = self.table.item(row, col)
                        row_data.append(item.text() if item else '')
                data.append(row_data)

            progress.setValue(60)
            QApplication.processEvents()

            # Добавляем строку с итогами
            total_row = ['ИТОГО:']
            for col in range(1, len(headers)):
                if headers[col] in ['Категория', 'Статус', 'Документы']:
                    col_data = [row[col] for row in data if row[col] and row[col].strip()]
                    total_row.append(str(len(col_data)))
                else:
                    total_row.append('')
            data.append(total_row)

            progress.setValue(80)
            QApplication.processEvents()

            # Создаем DataFrame и сохраняем
            df = pd.DataFrame(data, columns=headers)
            df.to_csv(file_path, index=False, encoding='utf-8-sig')

            progress.setValue(100)
            return True

        except Exception as e:
            print(f"Ошибка экспорта в CSV: {e}")
            return False

    # Добавляем настройки по умолчанию
    def init_settings(self):
        """Инициализация настроек по умолчанию"""
        cursor = self.conn.cursor()

        # Дни недели для работы (пн-пт = 1,2,3,4,5, сб,вс = 0)
        default_days = '1,2,3,4,5'  # пн-пт
        cursor.execute('''
            INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)
        ''', ('work_days', default_days))

        self.conn.commit()

    def get_work_days(self):
        """Получение дней недели для работы"""
        cursor = self.conn.cursor()
        cursor.execute('SELECT value FROM settings WHERE key = ?', ('work_days',))
        result = cursor.fetchone()
        if result:
            return [int(d) for d in result['value'].split(',')]
        return [1, 2, 3, 4, 5]  # по умолчанию пн-пт

    def set_work_days(self, days):
        """Установка дней недели для работы"""
        cursor = self.conn.cursor()
        days_str = ','.join([str(d) for d in days])
        cursor.execute('''
            INSERT OR REPLACE INTO settings (key, value, updated_at) VALUES (?, ?, CURRENT_TIMESTAMP)
        ''', ('work_days', days_str))
        self.conn.commit()

    def closeEvent(self, event):
        """Обработка закрытия окна"""
        self.db.close()
        event.accept()


class Application:
    """Главный класс приложения"""

    def __init__(self):
        self.app = QApplication(sys.argv)
        self.set_styles()

        # Показываем окно авторизации
        self.show_login()

    def set_styles(self):
        """Установка стилей приложения"""
        self.app.setStyle('Fusion')

        # Кастомная палитра
        palette = self.app.palette()
        palette.setColor(QPalette.ColorRole.Window, QColor(240, 240, 240))
        palette.setColor(QPalette.ColorRole.WindowText, QColor(50, 50, 50))
        palette.setColor(QPalette.ColorRole.Base, QColor(255, 255, 255))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(245, 245, 245))
        palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(255, 255, 255))
        palette.setColor(QPalette.ColorRole.ToolTipText, QColor(50, 50, 50))
        palette.setColor(QPalette.ColorRole.Text, QColor(50, 50, 50))
        palette.setColor(QPalette.ColorRole.Button, QColor(240, 240, 240))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor(50, 50, 50))
        palette.setColor(QPalette.ColorRole.Highlight, QColor(41, 128, 185))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
        self.app.setPalette(palette)

        # Стили для QPushButton
        self.app.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #1c6ea4;
            }
            QLineEdit, QComboBox, QTextEdit {
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 5px;
                background-color: white;
            }
            QLineEdit:focus, QComboBox:focus, QTextEdit:focus {
                border: 1px solid #3498db;
            }
            QTableWidget {
                background-color: white;
                alternate-background-color: #f9f9f9;
                gridline-color: #ddd;
                border: 1px solid #ddd;
            }
            QHeaderView::section {
                background-color: #3498db;
                color: white;
                padding: 5px;
                border: 1px solid #2980b9;
                font-weight: bold;
            }
            QTabWidget::pane {
                border: 1px solid #ddd;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #ecf0f1;
                padding: 8px 16px;
                margin-right: 2px;
                border: 1px solid #ddd;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #3498db;
                color: white;
            }
            QTabBar::tab:hover {
                background-color: #bdc3c7;
            }
        """)

    def show_login(self):
        """Показать окно авторизации"""
        self.login_window = LoginWindow()
        self.login_window.login_success.connect(self.on_login_success)
        self.login_window.show()

    def on_login_success(self, user_data):
        """Обработка успешной авторизации"""
        self.login_window.hide()
        self.main_window = MainWindow(user_data)
        self.main_window.logout_requested.connect(self.on_logout)
        self.main_window.show()

    def on_logout(self):
        """Обработка выхода из системы"""
        if hasattr(self, 'main_window'):
            self.main_window.close()
            self.main_window = None

        self.show_login()

    def run(self):
        """Запуск приложения"""
        return self.app.exec()


if __name__ == '__main__':
    application = Application()
    sys.exit(application.run())